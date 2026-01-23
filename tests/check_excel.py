# -*- coding: utf-8 -*-
"""
Excel 파일 검증 스크립트

생성된 Excel 파일의 설정을 확인합니다.
"""

import sys
from pathlib import Path

# 프로젝트 루트 경로 추가
_project_root = Path(__file__).parent.parent
sys.path.insert(0, str(_project_root))

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from config import EXCEL_MODULE_DIR


def check_excel(file_path: Path):
    """Excel 파일 설정 확인"""
    wb = load_workbook(file_path)
    ws = wb.active

    print(f'=== {file_path.name} 설정 ===')
    print(f'페이지 방향: {ws.page_setup.orientation}')
    print(f'용지 크기: {ws.page_setup.paperSize}')

    # 여백 (인치)
    pm = ws.page_margins
    print(f'여백: 좌{pm.left:.2f}in 우{pm.right:.2f}in 상{pm.top:.2f}in 하{pm.bottom:.2f}in')

    # 열 너비 합계
    total_width = 0
    col_count = 0
    for col in range(1, 50):
        w = ws.column_dimensions[get_column_letter(col)].width
        if w:
            total_width += w
            col_count += 1
            print(f'  열{col}: {w:.2f}')

    print(f'열 개수: {col_count}, 열 너비 합계: {total_width:.2f} 문자')

    # 행 높이 몇 개
    print('\n행 높이 (처음 5개):')
    for row in range(1, 6):
        h = ws.row_dimensions[row].height or 15
        print(f'  행{row}: {h:.2f}pt')


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='Excel 파일 설정 확인')
    parser.add_argument('file', nargs='?', default=None, help='Excel 파일 경로')
    args = parser.parse_args()

    if args.file:
        file_path = Path(args.file)
    else:
        # 기본값: excel/output.xlsx
        file_path = EXCEL_MODULE_DIR / 'output.xlsx'

    if file_path.exists():
        check_excel(file_path)
    else:
        print(f'파일을 찾을 수 없습니다: {file_path}')
