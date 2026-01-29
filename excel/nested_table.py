# -*- coding: utf-8 -*-
"""Nested 테이블 처리 모듈"""

import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import List, Dict, Union, Set, Tuple
from dataclasses import dataclass

# 프로젝트 루트 경로 설정
_project_root = Path(__file__).parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

from hwpxml.get_table_property import TableProperty
from hwpxml.get_cell_detail import CellDetail


@dataclass
class TableHierarchy:
    """테이블 계층 정보"""
    tbl_idx: int
    table_id: str
    parent_tbl_idx: int = -1  # -1이면 최상위 테이블
    parent_row: int = -1
    parent_col: int = -1
    row_count: int = 0
    col_count: int = 0


@dataclass
class CellPositionMapping:
    """셀 위치 매핑 정보 (원본 -> Excel)"""
    tbl_idx: int
    table_id: str
    orig_row: int
    orig_col: int
    excel_row: int  # 1-based
    excel_col: int  # 1-based
    list_id: str = ""
    is_nested: bool = False
    parent_tbl_idx: int = -1
    sheet_name: str = ""  # Excel 시트 이름


class NestedTableHandler:
    """Nested 테이블 처리 클래스"""

    # 엑셀 열 너비 변환 계수
    HWPUNIT_TO_EXCEL_WIDTH = 632
    # 엑셀 행 높이는 포인트 단위
    HWPUNIT_TO_PT = 100

    def __init__(self):
        pass

    def parse_table_hierarchy(self, hwpx_path: Union[str, Path]) -> List[TableHierarchy]:
        """HWPX에서 테이블 계층 구조 파악"""
        hwpx_path = Path(hwpx_path)
        hierarchy = []
        tbl_idx = 0

        with zipfile.ZipFile(hwpx_path, 'r') as zf:
            # section 파일 목록
            section_files = sorted([
                f for f in zf.namelist()
                if f.startswith('Contents/section') and f.endswith('.xml')
            ])

            for section_file in section_files:
                xml_content = zf.read(section_file)
                root = ET.fromstring(xml_content)

                # 재귀적으로 테이블 파싱
                tbl_idx = self._parse_tables_recursive(
                    root, hierarchy, tbl_idx, parent_tbl_idx=-1, parent_row=-1, parent_col=-1
                )

        return hierarchy

    def _parse_tables_recursive(
        self, element: ET.Element, hierarchy: List[TableHierarchy],
        tbl_idx: int, parent_tbl_idx: int, parent_row: int, parent_col: int
    ) -> int:
        """재귀적으로 테이블 파싱하여 계층 구조 수집"""
        for child in element:
            if child.tag.endswith('}tbl'):
                # 테이블 정보 추출
                table_id = child.get('id', '')
                row_cnt = int(child.get('rowCnt', 0))
                col_cnt = int(child.get('colCnt', 0))

                hierarchy.append(TableHierarchy(
                    tbl_idx=tbl_idx,
                    table_id=table_id,
                    parent_tbl_idx=parent_tbl_idx,
                    parent_row=parent_row,
                    parent_col=parent_col,
                    row_count=row_cnt,
                    col_count=col_cnt
                ))

                current_tbl_idx = tbl_idx
                tbl_idx += 1

                # 하위 셀에서 중첩 테이블 찾기
                for tr in child:
                    if not tr.tag.endswith('}tr'):
                        continue
                    for tc in tr:
                        if not tc.tag.endswith('}tc'):
                            continue
                        # 셀 위치 추출
                        cell_row, cell_col = 0, 0
                        for cell_child in tc:
                            if cell_child.tag.endswith('}cellAddr'):
                                cell_row = int(cell_child.get('rowAddr', 0))
                                cell_col = int(cell_child.get('colAddr', 0))
                                break
                        # 셀 내 중첩 테이블 재귀 탐색
                        tbl_idx = self._parse_tables_recursive(
                            tc, hierarchy, tbl_idx, current_tbl_idx, cell_row, cell_col
                        )
            else:
                # 다른 요소 내부도 탐색 (tbl이 아닌 경우)
                tbl_idx = self._parse_tables_recursive(
                    child, hierarchy, tbl_idx, parent_tbl_idx, parent_row, parent_col
                )

        return tbl_idx

    def calculate_inline_nested_expansion(
        self,
        parent_table: TableProperty,
        parent_cell_details: List[CellDetail],
        nested_tables_for_parent: List[TableHierarchy],
        all_tables: List[TableProperty]
    ) -> Dict:
        """
        inline nested 모드에서 부모 테이블의 행/열 확장 정보 계산

        Args:
            parent_table: 부모 테이블 정보
            parent_cell_details: 부모 테이블의 셀 디테일
            nested_tables_for_parent: 이 부모 테이블에 속한 nested 테이블 목록
            all_tables: 모든 테이블 정보

        Returns:
            {
                'row_expansion': {orig_row: extra_rows, ...},  # 각 행에 추가될 행 수
                'col_expansion': {orig_col: extra_cols, ...},  # 각 열에 추가될 열 수
                'nested_positions': [(nested_tbl_idx, parent_row, parent_col, nested_table), ...]
            }
        """
        row_expansion = {}  # 원본 행 -> 추가 행 수
        col_expansion = {}  # 원본 열 -> 추가 열 수

        # 부모 셀의 row_span, col_span 정보 맵
        cell_span_map = {}
        for cd in parent_cell_details:
            cell_span_map[(cd.row, cd.col)] = (cd.row_span, cd.col_span)

        nested_positions = []

        for nested_h in nested_tables_for_parent:
            nested_tbl_idx = nested_h.tbl_idx
            parent_row = nested_h.parent_row
            parent_col = nested_h.parent_col

            if nested_tbl_idx >= len(all_tables):
                continue

            nested_table = all_tables[nested_tbl_idx]
            nested_positions.append((nested_tbl_idx, parent_row, parent_col, nested_table))

            # 부모 셀의 span 정보 가져오기
            parent_row_span, parent_col_span = cell_span_map.get((parent_row, parent_col), (1, 1))

            # nested 테이블 크기
            nested_rows = nested_h.row_count
            nested_cols = nested_h.col_count

            # 부모 셀이 차지하는 영역보다 nested가 더 크면 확장 필요
            # 행 확장: nested_rows가 parent_row_span보다 크면 (nested_rows - parent_row_span)만큼 추가
            if nested_rows > parent_row_span:
                extra_rows = nested_rows - parent_row_span
                # 해당 행에 이미 확장이 있으면 최대값 사용
                current_extra = row_expansion.get(parent_row, 0)
                row_expansion[parent_row] = max(current_extra, extra_rows)

            # 열 확장: nested_cols가 parent_col_span보다 크면 (nested_cols - parent_col_span)만큼 추가
            if nested_cols > parent_col_span:
                extra_cols = nested_cols - parent_col_span
                current_extra = col_expansion.get(parent_col, 0)
                col_expansion[parent_col] = max(current_extra, extra_cols)

        return {
            'row_expansion': row_expansion,
            'col_expansion': col_expansion,
            'nested_positions': nested_positions
        }

    def build_expanded_row_mapping(
        self,
        parent_row_count: int,
        row_expansion: Dict[int, int]
    ) -> Dict[int, int]:
        """
        원본 행 -> 확장된 시작 행 매핑 생성

        Returns:
            {orig_row: new_start_row_offset, ...}
        """
        mapping = {}
        cumulative = 0
        for row in range(parent_row_count):
            mapping[row] = cumulative
            # 기본 1행 + 추가 확장 행
            extra = row_expansion.get(row, 0)
            cumulative += 1 + extra
        return mapping

    def build_expanded_col_mapping(
        self,
        parent_col_count: int,
        col_expansion: Dict[int, int]
    ) -> Dict[int, int]:
        """
        원본 열 -> 확장된 시작 열 매핑 생성

        Returns:
            {orig_col: new_start_col_offset, ...}
        """
        mapping = {}
        cumulative = 0
        for col in range(parent_col_count):
            mapping[col] = cumulative
            extra = col_expansion.get(col, 0)
            cumulative += 1 + extra
        return mapping

    def place_table_with_inline_nested(
        self,
        ws: Worksheet,
        parent_table: TableProperty,
        parent_cell_details: List[CellDetail],
        nested_tables_info: List[TableHierarchy],
        all_tables: List[TableProperty],
        all_cell_details: List[List[CellDetail]],
        start_row: int = 1,
        parent_tbl_idx: int = -1,
        get_column_widths_func=None,
        get_row_heights_func=None,
        apply_cell_style_func=None,
        apply_merged_cell_borders_func=None,
        hwp_color_to_rgb_func=None
    ) -> Tuple[int, List[CellPositionMapping]]:
        """
        부모 테이블에 nested table을 인라인으로 배치

        Args:
            ws: 워크시트
            parent_table: 부모 테이블 정보
            parent_cell_details: 부모 테이블의 셀 디테일
            nested_tables_info: 이 테이블에 속한 nested tables 계층 정보
            all_tables: 모든 테이블 정보
            all_cell_details: 모든 테이블의 셀 디테일
            start_row: 시작 행 (1-based)
            parent_tbl_idx: 부모 테이블 인덱스 (메타 정보용)
            get_column_widths_func: 열 너비 가져오기 함수 (외부 의존성)
            get_row_heights_func: 행 높이 가져오기 함수 (외부 의존성)
            apply_cell_style_func: 셀 스타일 적용 함수 (외부 의존성)
            apply_merged_cell_borders_func: 병합 셀 테두리 적용 함수 (외부 의존성)
            hwp_color_to_rgb_func: HWP 색상 변환 함수 (외부 의존성)

        Returns:
            (사용된 행 수, 셀 위치 매핑 리스트)
        """
        # 1. 확장 정보 계산
        expansion_info = self.calculate_inline_nested_expansion(
            parent_table, parent_cell_details, nested_tables_info, all_tables
        )
        row_expansion = expansion_info['row_expansion']
        col_expansion = expansion_info['col_expansion']
        nested_positions = expansion_info['nested_positions']

        # 2. 행/열 매핑 생성
        row_mapping = self.build_expanded_row_mapping(parent_table.row_count, row_expansion)
        col_mapping = self.build_expanded_col_mapping(parent_table.col_count, col_expansion)

        # 3. 확장된 총 행/열 수 계산
        total_rows = sum(1 + row_expansion.get(r, 0) for r in range(parent_table.row_count))
        total_cols = sum(1 + col_expansion.get(c, 0) for c in range(parent_table.col_count))

        # 4. 확장된 열 너비 설정
        if get_column_widths_func:
            parent_col_widths = get_column_widths_func(parent_table)
            expanded_col_idx = 1
            for orig_col in range(parent_table.col_count):
                orig_width = parent_col_widths[orig_col] if orig_col < len(parent_col_widths) else 5000
                extra_cols = col_expansion.get(orig_col, 0)

                # nested가 있는 열: nested 테이블의 열 너비 사용
                if extra_cols > 0:
                    # 해당 위치의 nested 테이블 찾기
                    nested_widths = None
                    for (nested_idx, p_row, p_col, nested_tbl) in nested_positions:
                        if p_col == orig_col:
                            nested_widths = get_column_widths_func(nested_tbl)
                            break

                    if nested_widths:
                        for i, nw in enumerate(nested_widths):
                            col_letter = get_column_letter(expanded_col_idx)
                            excel_width = max(nw / self.HWPUNIT_TO_EXCEL_WIDTH, 1)
                            ws.column_dimensions[col_letter].width = excel_width
                            expanded_col_idx += 1
                    else:
                        # nested 열 너비 정보 없으면 균등 분배
                        per_col_width = orig_width // (1 + extra_cols)
                        for _ in range(1 + extra_cols):
                            col_letter = get_column_letter(expanded_col_idx)
                            excel_width = max(per_col_width / self.HWPUNIT_TO_EXCEL_WIDTH, 1)
                            ws.column_dimensions[col_letter].width = excel_width
                            expanded_col_idx += 1
                else:
                    # 일반 열
                    col_letter = get_column_letter(expanded_col_idx)
                    excel_width = max(orig_width / self.HWPUNIT_TO_EXCEL_WIDTH, 1)
                    ws.column_dimensions[col_letter].width = excel_width
                    expanded_col_idx += 1

        # 5. 확장된 행 높이 설정
        if get_row_heights_func:
            parent_row_heights = get_row_heights_func(parent_table)
            expanded_row_idx = start_row
            for orig_row in range(parent_table.row_count):
                orig_height = parent_row_heights[orig_row] if orig_row < len(parent_row_heights) else 1500
                extra_rows = row_expansion.get(orig_row, 0)

                if extra_rows > 0:
                    # nested가 있는 행: nested 테이블의 행 높이 사용
                    nested_heights = None
                    for (nested_idx, p_row, p_col, nested_tbl) in nested_positions:
                        if p_row == orig_row:
                            nested_heights = get_row_heights_func(nested_tbl)
                            break

                    if nested_heights:
                        for nh in nested_heights:
                            height_pt = max(nh / self.HWPUNIT_TO_PT, 10)
                            ws.row_dimensions[expanded_row_idx].height = height_pt
                            expanded_row_idx += 1
                    else:
                        # nested 행 높이 정보 없으면 균등 분배
                        per_row_height = orig_height // (1 + extra_rows)
                        for _ in range(1 + extra_rows):
                            height_pt = max(per_row_height / self.HWPUNIT_TO_PT, 10)
                            ws.row_dimensions[expanded_row_idx].height = height_pt
                            expanded_row_idx += 1
                else:
                    # 일반 행
                    height_pt = max(orig_height / self.HWPUNIT_TO_PT, 10)
                    ws.row_dimensions[expanded_row_idx].height = height_pt
                    expanded_row_idx += 1

        # 6. nested 테이블이 있는 셀 위치 집합
        nested_cell_positions: Set[Tuple[int, int]] = set()
        for (nested_idx, p_row, p_col, nested_tbl) in nested_positions:
            nested_cell_positions.add((p_row, p_col))

        # 셀 위치 매핑 리스트 초기화
        cell_mappings: List[CellPositionMapping] = []
        parent_table_id = str(parent_table.id) if parent_table.id else ""

        # 7. 부모 테이블 셀 배치 (nested 위치 제외)
        for cd in parent_cell_details:
            orig_row, orig_col = cd.row, cd.col

            # nested 테이블이 있는 셀은 나중에 처리
            if (orig_row, orig_col) in nested_cell_positions:
                continue

            # 확장된 위치 계산
            new_row = start_row + row_mapping[orig_row]
            new_col = col_mapping[orig_col] + 1  # 1-based

            # 부모 셀 위치 매핑 추가
            cell_mappings.append(CellPositionMapping(
                tbl_idx=parent_tbl_idx,
                table_id=parent_table_id,
                orig_row=orig_row,
                orig_col=orig_col,
                excel_row=new_row,
                excel_col=new_col,
                list_id=cd.list_id or "",
                is_nested=False,
                parent_tbl_idx=-1
            ))

            # 확장된 row_span, col_span 계산
            new_row_span = 0
            for r in range(cd.row_span):
                if orig_row + r < parent_table.row_count:
                    new_row_span += 1 + row_expansion.get(orig_row + r, 0)

            new_col_span = 0
            for c in range(cd.col_span):
                if orig_col + c < parent_table.col_count:
                    new_col_span += 1 + col_expansion.get(orig_col + c, 0)

            # 셀 값 및 스타일 적용
            try:
                excel_cell = ws.cell(row=new_row, column=new_col)
                excel_cell.value = cd.text
                if apply_cell_style_func:
                    apply_cell_style_func(excel_cell, cd, 0, 1)

                # 병합 처리
                if new_row_span > 1 or new_col_span > 1:
                    try:
                        ws.merge_cells(
                            start_row=new_row,
                            start_column=new_col,
                            end_row=new_row + new_row_span - 1,
                            end_column=new_col + new_col_span - 1
                        )
                        # 병합 영역 테두리 적용
                        if apply_merged_cell_borders_func:
                            apply_merged_cell_borders_func(
                                ws, cd, new_row, new_col,
                                new_row + new_row_span - 1, new_col + new_col_span - 1
                            )
                    except ValueError:
                        pass
            except Exception:
                pass

        # 8. nested 테이블 인라인 배치 (점선 테두리) + 부모 셀 외곽 테두리
        # nested_tables_info에서 테이블 정보 매핑 생성
        nested_info_map = {h.tbl_idx: h for h in nested_tables_info}

        for (nested_idx, p_row, p_col, nested_tbl) in nested_positions:
            if nested_idx >= len(all_cell_details):
                continue

            nested_cell_details = all_cell_details[nested_idx]
            nested_hierarchy = nested_info_map.get(nested_idx)
            nested_table_id = nested_hierarchy.table_id if nested_hierarchy else f"tbl_{nested_idx}"

            # 부모 셀의 확장된 시작 위치
            parent_new_row = start_row + row_mapping[p_row]
            parent_new_col = col_mapping[p_col] + 1  # 1-based

            # nested 테이블의 행/열 수
            nested_row_count = nested_tbl.row_count
            nested_col_count = nested_tbl.col_count

            # 부모 셀의 CellDetail 찾기 (테두리 정보용)
            parent_cd = None
            for cd in parent_cell_details:
                if cd.row == p_row and cd.col == p_col:
                    parent_cd = cd
                    break

            # nested 테이블 셀 배치
            for ncd in nested_cell_details:
                nested_row = parent_new_row + ncd.row
                nested_col = parent_new_col + ncd.col

                # nested 셀 위치 매핑 추가
                cell_mappings.append(CellPositionMapping(
                    tbl_idx=nested_idx,
                    table_id=nested_table_id,
                    orig_row=ncd.row,
                    orig_col=ncd.col,
                    excel_row=nested_row,
                    excel_col=nested_col,
                    list_id=ncd.list_id or "",
                    is_nested=True,
                    parent_tbl_idx=parent_tbl_idx
                ))

                # 셀이 nested 테이블 내에서의 위치 (외곽 판단용)
                is_first_row = (ncd.row == 0)
                is_last_row = (ncd.row + ncd.row_span - 1 == nested_row_count - 1)
                is_first_col = (ncd.col == 0)
                is_last_col = (ncd.col + ncd.col_span - 1 == nested_col_count - 1)

                try:
                    excel_cell = ws.cell(row=nested_row, column=nested_col)
                    excel_cell.value = ncd.text

                    # 스타일 적용 (위치 기반 테두리)
                    self._apply_nested_cell_style_with_position(
                        excel_cell, ncd,
                        is_first_row, is_last_row, is_first_col, is_last_col,
                        parent_cd, hwp_color_to_rgb_func
                    )

                    # 병합 처리
                    if ncd.row_span > 1 or ncd.col_span > 1:
                        try:
                            ws.merge_cells(
                                start_row=nested_row,
                                start_column=nested_col,
                                end_row=nested_row + ncd.row_span - 1,
                                end_column=nested_col + ncd.col_span - 1
                            )
                            # 병합 영역에도 위치 기반 테두리
                            self._apply_nested_merged_borders_with_position(
                                ws, ncd, nested_row, nested_col,
                                nested_row + ncd.row_span - 1, nested_col + ncd.col_span - 1,
                                is_first_row, is_last_row, is_first_col, is_last_col,
                                parent_cd
                            )
                        except ValueError:
                            pass
                except Exception:
                    pass

        return total_rows, cell_mappings

    def _apply_nested_cell_style(
        self, excel_cell, cd: CellDetail, para_idx: int, total_paras: int,
        hwp_color_to_rgb_func=None
    ):
        """
        nested 테이블 셀에 스타일 적용 (점선 테두리)
        """
        # 점선 테두리 스타일
        dashed_side = Side(style='dashed', color='808080')

        border = Border(
            left=dashed_side,
            right=dashed_side,
            top=dashed_side if para_idx == 0 else Side(),
            bottom=dashed_side if para_idx == total_paras - 1 else Side(),
        )
        excel_cell.border = border

        # 배경색
        if hwp_color_to_rgb_func:
            bg_color = hwp_color_to_rgb_func(cd.border.bg_color)
            if bg_color and bg_color != 'FFFFFF':
                excel_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        # 폰트
        para_font = cd.font
        if para_idx < len(cd.paragraphs) and cd.paragraphs[para_idx].font:
            para_font = cd.paragraphs[para_idx].font

        font_color = None
        if hwp_color_to_rgb_func:
            font_color = hwp_color_to_rgb_func(para_font.color)

        excel_cell.font = Font(
            name=para_font.name if para_font.name else None,
            size=para_font.size_pt() if para_font.size > 0 else None,
            bold=para_font.bold,
            italic=para_font.italic,
            underline='single' if para_font.underline else None,
            strike=para_font.strikeout,
            color=font_color if font_color else None,
        )

        # 정렬
        h_align_map = {'LEFT': 'left', 'CENTER': 'center', 'RIGHT': 'right', 'JUSTIFY': 'justify'}
        v_align_map = {'TOP': 'top', 'CENTER': 'center', 'BOTTOM': 'bottom', 'BASELINE': 'center'}
        h_align = 'left'
        v_align = 'center'
        if cd.paragraphs:
            h_align = h_align_map.get(cd.paragraphs[0].align_h, 'left')
            v_align = v_align_map.get(cd.paragraphs[0].align_v, 'center')
        excel_cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=True)

    def _apply_nested_merged_borders(
        self, ws, cd: CellDetail,
        start_row: int, start_col: int, end_row: int, end_col: int
    ):
        """nested 테이블의 병합 셀 영역에 점선 테두리 적용"""
        if start_row == end_row and start_col == end_col:
            return

        dashed_side = Side(style='dashed', color='808080')

        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                try:
                    cell = ws.cell(row=r, column=c)
                    border = Border(
                        left=dashed_side if c == start_col else Side(),
                        right=dashed_side if c == end_col else Side(),
                        top=dashed_side if r == start_row else Side(),
                        bottom=dashed_side if r == end_row else Side(),
                    )
                    cell.border = border
                except:
                    pass

    def _apply_nested_cell_style_with_position(
        self, excel_cell, cd: CellDetail,
        is_first_row: bool, is_last_row: bool,
        is_first_col: bool, is_last_col: bool,
        parent_cd: CellDetail = None,
        hwp_color_to_rgb_func=None
    ):
        """
        nested 테이블 셀에 위치 기반 스타일 적용
        - 내부 셀: 점선 테두리
        - 외곽 셀: 부모 셀의 테두리 스타일 상속 (실선)
        """
        # 점선 테두리 (내부용)
        dashed_side = Side(style='dashed', color='808080')
        # 실선 테두리 (외곽용 - 부모 셀 테두리 상속)
        solid_side = Side(style='thin', color='000000')

        # 부모 셀의 테두리 정보가 있으면 사용
        if parent_cd and parent_cd.border:
            parent_border = parent_cd.border
            # 부모 테두리 타입을 Side로 변환
            top_side = self._get_border_side_from_type(parent_border.top) if is_first_row else dashed_side
            bottom_side = self._get_border_side_from_type(parent_border.bottom) if is_last_row else dashed_side
            left_side = self._get_border_side_from_type(parent_border.left) if is_first_col else dashed_side
            right_side = self._get_border_side_from_type(parent_border.right) if is_last_col else dashed_side
        else:
            # 부모 테두리 정보 없으면 기본값
            top_side = solid_side if is_first_row else dashed_side
            bottom_side = solid_side if is_last_row else dashed_side
            left_side = solid_side if is_first_col else dashed_side
            right_side = solid_side if is_last_col else dashed_side

        border = Border(
            left=left_side,
            right=right_side,
            top=top_side,
            bottom=bottom_side,
        )
        excel_cell.border = border

        # 배경색
        if hwp_color_to_rgb_func:
            bg_color = hwp_color_to_rgb_func(cd.border.bg_color)
            if bg_color and bg_color != 'FFFFFF':
                excel_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        # 폰트
        para_font = cd.font
        if cd.paragraphs and cd.paragraphs[0].font:
            para_font = cd.paragraphs[0].font

        font_color = None
        if hwp_color_to_rgb_func:
            font_color = hwp_color_to_rgb_func(para_font.color)

        excel_cell.font = Font(
            name=para_font.name if para_font.name else None,
            size=para_font.size_pt() if para_font.size > 0 else None,
            bold=para_font.bold,
            italic=para_font.italic,
            underline='single' if para_font.underline else None,
            strike=para_font.strikeout,
            color=font_color if font_color else None,
        )

        # 정렬
        h_align_map = {'LEFT': 'left', 'CENTER': 'center', 'RIGHT': 'right', 'JUSTIFY': 'justify'}
        v_align_map = {'TOP': 'top', 'CENTER': 'center', 'BOTTOM': 'bottom', 'BASELINE': 'center'}
        h_align = 'left'
        v_align = 'center'
        if cd.paragraphs:
            h_align = h_align_map.get(cd.paragraphs[0].align_h, 'left')
            v_align = v_align_map.get(cd.paragraphs[0].align_v, 'center')
        excel_cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=True)

    def _apply_nested_merged_borders_with_position(
        self, ws, cd: CellDetail,
        start_row: int, start_col: int, end_row: int, end_col: int,
        is_first_row: bool, is_last_row: bool,
        is_first_col: bool, is_last_col: bool,
        parent_cd: CellDetail = None
    ):
        """nested 테이블의 병합 셀 영역에 위치 기반 테두리 적용"""
        if start_row == end_row and start_col == end_col:
            return

        dashed_side = Side(style='dashed', color='808080')
        solid_side = Side(style='thin', color='000000')

        # 부모 셀의 테두리 정보
        if parent_cd and parent_cd.border:
            parent_border = parent_cd.border
            outer_top = self._get_border_side_from_type(parent_border.top) if is_first_row else dashed_side
            outer_bottom = self._get_border_side_from_type(parent_border.bottom) if is_last_row else dashed_side
            outer_left = self._get_border_side_from_type(parent_border.left) if is_first_col else dashed_side
            outer_right = self._get_border_side_from_type(parent_border.right) if is_last_col else dashed_side
        else:
            outer_top = solid_side if is_first_row else dashed_side
            outer_bottom = solid_side if is_last_row else dashed_side
            outer_left = solid_side if is_first_col else dashed_side
            outer_right = solid_side if is_last_col else dashed_side

        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                try:
                    cell = ws.cell(row=r, column=c)

                    # 병합 영역의 각 위치에 따른 테두리
                    left = outer_left if c == start_col else Side()
                    right = outer_right if c == end_col else Side()
                    top = outer_top if r == start_row else Side()
                    bottom = outer_bottom if r == end_row else Side()

                    border = Border(left=left, right=right, top=top, bottom=bottom)
                    cell.border = border
                except:
                    pass

    def _get_border_side_from_type(self, border_type: str) -> Side:
        """HWP 테두리 타입을 openpyxl Side로 변환"""
        border_style_map = {
            'NONE': None,
            'SOLID': 'thin',
            'DOUBLE': 'double',
            'DOTTED': 'dotted',
            'DASHED': 'dashed',
            'DASH_DOT': 'dashDot',
            'DASH_DOT_DOT': 'dashDotDot',
            'THICK': 'medium',
            'THICK_DOUBLE': 'double',
            'THICK_DOUBLE_SLIM': 'double',
        }
        style = border_style_map.get(border_type, 'thin')
        if style is None:
            return Side()
        return Side(style=style, color='000000')
