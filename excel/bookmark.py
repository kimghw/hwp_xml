# -*- coding: utf-8 -*-
"""북마크 기반 테이블 추출 모듈"""

import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import List, Dict, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties


class BookmarkHandler:
    """북마크 처리 클래스"""

    def __init__(self, converter):
        """
        Args:
            converter: HwpxToExcel 인스턴스 (테이블 파싱/변환 기능 사용)
        """
        self.converter = converter

    def get_bookmarks(self, hwpx_path: Union[str, Path]) -> List[dict]:
        """
        HWPX 파일에서 북마크 목록 추출

        Returns:
            [{"name": "북마크명", "section": 섹션인덱스, "position": XML내위치}, ...]
        """
        hwpx_path = Path(hwpx_path)
        bookmarks = []

        with zipfile.ZipFile(hwpx_path, 'r') as zf:
            section_files = sorted([
                f for f in zf.namelist()
                if f.startswith('Contents/section') and f.endswith('.xml')
            ])

            for section_idx, section_file in enumerate(section_files):
                xml_content = zf.read(section_file)
                root = ET.fromstring(xml_content)

                # 모든 bookmark 태그 찾기
                for elem in root.iter():
                    if elem.tag.endswith('}bookmark'):
                        name = elem.get('name', '')
                        if name:
                            bookmarks.append({
                                "name": name,
                                "section": section_idx,
                                "element": elem  # 위치 추적용
                            })

        return bookmarks

    def get_bookmark_table_mapping(self, hwpx_path: Union[str, Path]) -> Dict[str, List[int]]:
        """
        북마크별로 소속 테이블 인덱스 매핑

        Returns:
            {"북마크명": [테이블인덱스1, 테이블인덱스2, ...], ...}
        """
        hwpx_path = Path(hwpx_path)
        result = {}

        with zipfile.ZipFile(hwpx_path, 'r') as zf:
            section_files = sorted([
                f for f in zf.namelist()
                if f.startswith('Contents/section') and f.endswith('.xml')
            ])

            for section_file in section_files:
                xml_content = zf.read(section_file)
                root = ET.fromstring(xml_content)

                # 문서 순서대로 북마크와 테이블 위치 추출
                current_bookmark = None
                table_idx = 0

                for elem in root.iter():
                    if elem.tag.endswith('}bookmark'):
                        name = elem.get('name', '')
                        if name:
                            current_bookmark = name
                            if current_bookmark not in result:
                                result[current_bookmark] = []
                    elif elem.tag.endswith('}tbl'):
                        # 중첩 테이블이 아닌 최상위 테이블만 카운트
                        # 부모가 tc가 아닌 경우
                        if current_bookmark:
                            result[current_bookmark].append(table_idx)
                        table_idx += 1

        return result

    def get_body_elements(self, hwpx_path: Union[str, Path]) -> List[dict]:
        """
        HWPX 파일에서 본문 요소(문단, 테이블) 순서대로 추출

        Returns:
            [{"type": "para"|"table", "text": str, "table_idx": int, ...}, ...]
        """
        hwpx_path = Path(hwpx_path)
        elements = []
        table_idx = 0

        with zipfile.ZipFile(hwpx_path, 'r') as zf:
            section_files = sorted([
                f for f in zf.namelist()
                if f.startswith('Contents/section') and f.endswith('.xml')
            ])

            for section_file in section_files:
                xml_content = zf.read(section_file)
                root = ET.fromstring(xml_content)

                # root의 직접 자식만 순회 (p 태그)
                for child in root:
                    tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag

                    if tag == 'p':
                        # 문단에서 텍스트 추출
                        texts = []
                        has_table = False

                        for elem in child.iter():
                            elem_tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                            if elem_tag == 't' and elem.text:
                                texts.append(elem.text)
                            elif elem_tag == 'tbl':
                                has_table = True

                        text = ''.join(texts).strip()

                        # caption 태그 제거 (테이블 캡션)
                        if text.startswith('{caption:'):
                            # {caption:tbl_0|} 형식에서 실제 텍스트만 추출
                            if '|}' in text:
                                text = text.split('|}', 1)[1].strip()
                            elif '|' in text:
                                text = text.split('|', 1)[1].strip()

                        if has_table:
                            # 테이블 포함 문단
                            elements.append({
                                "type": "table",
                                "table_idx": table_idx,
                                "caption": text
                            })
                            table_idx += 1
                        elif text:
                            # 일반 본문 문단
                            elements.append({
                                "type": "para",
                                "text": text
                            })

        return elements

    def get_bookmark_body_mapping(self, hwpx_path: Union[str, Path]) -> Dict[str, List[dict]]:
        """
        북마크별로 본문 요소(문단, 테이블) 매핑

        Returns:
            {"북마크명": [{"type": "para"|"table", ...}, ...], ...}
        """
        hwpx_path = Path(hwpx_path)
        result = {}
        current_bookmark = None
        table_idx = 0

        with zipfile.ZipFile(hwpx_path, 'r') as zf:
            section_files = sorted([
                f for f in zf.namelist()
                if f.startswith('Contents/section') and f.endswith('.xml')
            ])

            for section_file in section_files:
                xml_content = zf.read(section_file)
                root = ET.fromstring(xml_content)

                for child in root:
                    tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag

                    if tag == 'p':
                        # 북마크 확인
                        for elem in child.iter():
                            elem_tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                            if elem_tag == 'bookmark':
                                name = elem.get('name', '')
                                if name:
                                    current_bookmark = name
                                    if current_bookmark not in result:
                                        result[current_bookmark] = []

                        # 텍스트 및 테이블 추출
                        texts = []
                        has_table = False

                        for elem in child.iter():
                            elem_tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                            if elem_tag == 't' and elem.text:
                                texts.append(elem.text)
                            elif elem_tag == 'tbl':
                                has_table = True

                        text = ''.join(texts).strip()

                        # caption 제거
                        if text.startswith('{caption:'):
                            if '|}' in text:
                                text = text.split('|}', 1)[1].strip()
                            elif '|' in text:
                                text = text.split('|', 1)[1].strip()

                        if current_bookmark:
                            if has_table:
                                result[current_bookmark].append({
                                    "type": "table",
                                    "table_idx": table_idx,
                                    "caption": text
                                })
                                table_idx += 1
                            elif text:
                                result[current_bookmark].append({
                                    "type": "para",
                                    "text": text
                                })

        return result

    def convert_by_bookmark(
        self,
        hwpx_path: Union[str, Path],
        bookmark_name: str,
        output_path: Optional[Union[str, Path]] = None,
        split_by_para: bool = False,
        include_cell_info: bool = False,
        hide_para_rows: bool = True
    ) -> Path:
        """
        특정 북마크 섹션의 테이블만 하나의 시트로 변환

        Args:
            hwpx_path: HWPX 파일 경로
            bookmark_name: 북마크 이름 (부분 일치 지원)
            output_path: 출력 Excel 경로
            split_by_para: 문단별 행 분할 여부
            include_cell_info: 셀 상세 정보 시트 포함 여부
            hide_para_rows: para_id 행 숨김 여부 (include_cell_info=True일 때)

        Returns:
            생성된 Excel 파일 경로
        """
        hwpx_path = Path(hwpx_path)

        # 북마크-테이블 매핑 가져오기
        bookmark_mapping = self.get_bookmark_table_mapping(hwpx_path)

        # 북마크 이름 찾기 (부분 일치)
        matched_bookmark = None
        for bm_name in bookmark_mapping.keys():
            if bookmark_name in bm_name:
                matched_bookmark = bm_name
                break

        if not matched_bookmark:
            raise ValueError(f"북마크를 찾을 수 없습니다: {bookmark_name}")

        table_indices = bookmark_mapping[matched_bookmark]
        if not table_indices:
            raise ValueError(f"북마크 '{matched_bookmark}'에 테이블이 없습니다")

        # 출력 경로 설정
        if output_path is None:
            safe_name = bookmark_name.replace(" ", "_").replace("/", "_")[:20]
            output_path = hwpx_path.with_name(f"{hwpx_path.stem}_{safe_name}.xlsx")
        else:
            output_path = Path(output_path)

        # 모든 테이블 데이터 로드
        all_tables = self.converter.table_parser.from_hwpx(hwpx_path)
        pages = self.converter.page_parser.from_hwpx(hwpx_path)
        all_cell_details = self.converter.cell_detail_parser.from_hwpx_by_table(hwpx_path)

        # 해당 북마크의 테이블만 필터링
        tables = [all_tables[i] for i in table_indices if i < len(all_tables)]
        cell_details_list = [all_cell_details[i] for i in table_indices if i < len(all_cell_details)]

        if not tables:
            raise ValueError(f"테이블 인덱스가 범위를 벗어났습니다: {table_indices}")

        page = pages[0] if pages else None

        # 통합 열 그리드 생성
        unified_col_boundaries = self.converter._build_unified_column_grid(tables)
        unified_col_widths = []
        for i in range(len(unified_col_boundaries) - 1):
            unified_col_widths.append(unified_col_boundaries[i + 1] - unified_col_boundaries[i])

        # Excel 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = matched_bookmark[:31] if len(matched_bookmark) <= 31 else matched_bookmark[:28] + "..."

        if page:
            self.converter._apply_page_settings(ws, page)

        # 통합 열 너비 적용
        for col_idx, width in enumerate(unified_col_widths, start=1):
            col_letter = get_column_letter(col_idx)
            excel_width = width / self.converter.HWPUNIT_TO_EXCEL_WIDTH
            excel_width = max(excel_width, 1)
            ws.column_dimensions[col_letter].width = excel_width

        # 각 테이블 배치
        current_row = 1
        for tbl_idx, table in enumerate(tables):
            table_col_boundaries = self.converter._get_table_column_boundaries(table)
            col_mapping = self.converter._map_columns_to_unified(table_col_boundaries, unified_col_boundaries)
            cell_details = cell_details_list[tbl_idx] if tbl_idx < len(cell_details_list) else []

            if split_by_para and cell_details:
                rows_used = self.converter._place_table_with_para_split_unified(
                    ws, table, cell_details, col_mapping, unified_col_widths, current_row
                )
            else:
                rows_used = self.converter._place_table_unified(
                    ws, table, cell_details, col_mapping, unified_col_widths, current_row
                )

            current_row += rows_used + 1  # 테이블 사이 빈 행

        # 모든 열을 1페이지 폭에 맞추기
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 높이는 제한 없음
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        # 셀 정보 시트 추가 (옵션)
        if include_cell_info:
            try:
                from .cell_info_sheet import add_cell_info_sheet
            except ImportError:
                from cell_info_sheet import add_cell_info_sheet
            page_id = page.section_id if page and page.section_id else "section_0"
            add_cell_info_sheet(wb, hwpx_path, f"{matched_bookmark[:20]}_메타", hide_para_rows, page_id, "")

        wb.save(output_path)
        print(f"  북마크 '{matched_bookmark}': {len(tables)}개 테이블 -> {output_path}")
        return output_path

    def convert_by_bookmark_with_body(
        self,
        hwpx_path: Union[str, Path],
        bookmark_name: str,
        output_path: Optional[Union[str, Path]] = None,
        split_by_para: bool = False,
        include_cell_info: bool = False,
        hide_para_rows: bool = True
    ) -> Path:
        """
        특정 북마크 섹션의 본문(문단+테이블)을 하나의 시트로 변환

        본문 문단은 첫 열에 배치하고, 테이블은 통합 열 그리드에 맞춰 배치합니다.

        Args:
            hwpx_path: HWPX 파일 경로
            bookmark_name: 북마크 이름 (부분 일치 지원)
            output_path: 출력 Excel 경로
            split_by_para: 문단별 행 분할 여부
            include_cell_info: 셀 상세 정보 시트 포함 여부
            hide_para_rows: para_id 행 숨김 여부 (include_cell_info=True일 때)

        Returns:
            생성된 Excel 파일 경로
        """
        hwpx_path = Path(hwpx_path)

        # 북마크-본문 매핑 가져오기
        bookmark_body_mapping = self.get_bookmark_body_mapping(hwpx_path)

        # 북마크 이름 찾기 (부분 일치)
        matched_bookmark = None
        for bm_name in bookmark_body_mapping.keys():
            if bookmark_name in bm_name:
                matched_bookmark = bm_name
                break

        if not matched_bookmark:
            raise ValueError(f"북마크를 찾을 수 없습니다: {bookmark_name}")

        body_elements = bookmark_body_mapping[matched_bookmark]
        if not body_elements:
            raise ValueError(f"북마크 '{matched_bookmark}'에 내용이 없습니다")

        # 출력 경로 설정
        if output_path is None:
            safe_name = bookmark_name.replace(" ", "_").replace("/", "_")[:20]
            output_path = hwpx_path.with_name(f"{hwpx_path.stem}_{safe_name}_body.xlsx")
        else:
            output_path = Path(output_path)

        # 테이블 인덱스 추출
        table_indices = [e["table_idx"] for e in body_elements if e["type"] == "table"]

        # 모든 테이블 데이터 로드
        all_tables = self.converter.table_parser.from_hwpx(hwpx_path)
        pages = self.converter.page_parser.from_hwpx(hwpx_path)
        all_cell_details = self.converter.cell_detail_parser.from_hwpx_by_table(hwpx_path)

        # 해당 북마크의 테이블만 필터링
        tables = [all_tables[i] for i in table_indices if i < len(all_tables)]
        cell_details_map = {i: all_cell_details[i] for i in table_indices if i < len(all_cell_details)}

        page = pages[0] if pages else None

        # 통합 열 그리드 생성 (테이블이 있을 경우)
        if tables:
            unified_col_boundaries = self.converter._build_unified_column_grid(tables)
            unified_col_widths = []
            for i in range(len(unified_col_boundaries) - 1):
                unified_col_widths.append(unified_col_boundaries[i + 1] - unified_col_boundaries[i])
        else:
            # 테이블이 없으면 기본 열 1개
            unified_col_boundaries = [0, 48000]  # A4 폭 정도
            unified_col_widths = [48000]

        # Excel 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = matched_bookmark[:31] if len(matched_bookmark) <= 31 else matched_bookmark[:28] + "..."

        if page:
            self.converter._apply_page_settings(ws, page)

        # 통합 열 너비 적용
        for col_idx, width in enumerate(unified_col_widths, start=1):
            col_letter = get_column_letter(col_idx)
            excel_width = width / self.converter.HWPUNIT_TO_EXCEL_WIDTH
            excel_width = max(excel_width, 1)
            ws.column_dimensions[col_letter].width = excel_width

        # 본문 요소 순서대로 배치
        current_row = 1
        for elem in body_elements:
            if elem["type"] == "para":
                # 본문 문단: 첫 열에 텍스트 배치, 전체 열 병합
                cell = ws.cell(row=current_row, column=1, value=elem["text"])
                cell.alignment = Alignment(wrap_text=True, vertical='top')

                # 전체 열 병합
                if len(unified_col_widths) > 1:
                    try:
                        ws.merge_cells(
                            start_row=current_row,
                            start_column=1,
                            end_row=current_row,
                            end_column=len(unified_col_widths)
                        )
                    except ValueError:
                        pass

                current_row += 1

            elif elem["type"] == "table":
                tbl_idx = elem["table_idx"]
                if tbl_idx >= len(all_tables):
                    continue

                table = all_tables[tbl_idx]
                table_col_boundaries = self.converter._get_table_column_boundaries(table)
                col_mapping = self.converter._map_columns_to_unified(table_col_boundaries, unified_col_boundaries)
                cell_details = cell_details_map.get(tbl_idx, [])

                if split_by_para and cell_details:
                    rows_used = self.converter._place_table_with_para_split_unified(
                        ws, table, cell_details, col_mapping, unified_col_widths, current_row
                    )
                else:
                    rows_used = self.converter._place_table_unified(
                        ws, table, cell_details, col_mapping, unified_col_widths, current_row
                    )

                current_row += rows_used + 1  # 테이블 후 빈 행

        # 모든 열을 1페이지 폭에 맞추기
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        # 셀 정보 시트 추가 (옵션)
        if include_cell_info:
            try:
                from .cell_info_sheet import add_cell_info_sheet
            except ImportError:
                from cell_info_sheet import add_cell_info_sheet
            page_id = page.section_id if page and page.section_id else "section_0"
            add_cell_info_sheet(wb, hwpx_path, f"{matched_bookmark[:20]}_메타", hide_para_rows, page_id, "")

        wb.save(output_path)
        para_count = sum(1 for e in body_elements if e["type"] == "para")
        table_count = sum(1 for e in body_elements if e["type"] == "table")
        print(f"  북마크 '{matched_bookmark}': 본문 {para_count}개 + 테이블 {table_count}개 -> {output_path}")
        return output_path

    def convert_all_by_bookmark(
        self,
        hwpx_path: Union[str, Path],
        output_path: Optional[Union[str, Path]] = None,
        include_body: bool = True,
        split_by_para: bool = False,
        include_cell_info: bool = False,
        hide_para_rows: bool = True,
        inline_nested: bool = False
    ) -> Path:
        """
        전체 문서를 북마크별로 시트 분리하여 변환

        각 북마크 섹션이 별도의 시트로 생성됩니다.

        Args:
            hwpx_path: HWPX 파일 경로
            output_path: 출력 Excel 경로
            include_body: 본문 문단 포함 여부 (False면 테이블만)
            split_by_para: 문단별 행 분할 여부
            include_cell_info: 셀 상세 정보 시트 포함 여부
            hide_para_rows: para_id 행 숨김 여부 (include_cell_info=True일 때)
            inline_nested: True면 nested 테이블을 부모 셀에 인라인 배치

        Returns:
            생성된 Excel 파일 경로
        """
        hwpx_path = Path(hwpx_path)

        if output_path is None:
            suffix = "_by_bookmark_body" if include_body else "_by_bookmark"
            if inline_nested:
                suffix += "_inline"
            output_path = hwpx_path.with_name(f"{hwpx_path.stem}{suffix}.xlsx")
        else:
            output_path = Path(output_path)

        # 데이터 로드
        bookmark_body_mapping = self.get_bookmark_body_mapping(hwpx_path)
        all_tables = self.converter.table_parser.from_hwpx(hwpx_path)
        pages = self.converter.page_parser.from_hwpx(hwpx_path)
        all_cell_details = self.converter.cell_detail_parser.from_hwpx_by_table(hwpx_path)
        page = pages[0] if pages else None

        # inline_nested 모드: 테이블 계층 구조 파악
        table_hierarchy = []
        nested_by_parent = {}  # parent_tbl_idx -> [nested1, nested2, ...]
        if inline_nested:
            table_hierarchy = self.converter._parse_table_hierarchy(hwpx_path)
            for h in table_hierarchy:
                if h.parent_tbl_idx != -1:
                    if h.parent_tbl_idx not in nested_by_parent:
                        nested_by_parent[h.parent_tbl_idx] = []
                    nested_by_parent[h.parent_tbl_idx].append(h)

        wb = Workbook()
        default_sheet = wb.active

        # inline_nested 모드: 전체 셀 위치 매핑 수집
        all_cell_mappings = []

        for bm_idx, (bm_name, body_elements) in enumerate(bookmark_body_mapping.items()):
            if not body_elements:
                continue

            # 시트 이름 (31자 제한, 중복 방지)
            sheet_name = bm_name[:31] if len(bm_name) <= 31 else bm_name[:28] + "..."
            if sheet_name in wb.sheetnames:
                sheet_name = f"{bm_idx}_{sheet_name}"[:31]

            ws = wb.create_sheet(title=sheet_name)

            # 테이블 인덱스 추출
            table_indices = [e["table_idx"] for e in body_elements if e["type"] == "table"]
            tables = [all_tables[i] for i in table_indices if i < len(all_tables)]
            cell_details_map = {i: all_cell_details[i] for i in table_indices if i < len(all_cell_details)}

            # 통합 열 그리드 생성
            if tables:
                unified_col_boundaries = self.converter._build_unified_column_grid(tables)
                unified_col_widths = [
                    unified_col_boundaries[i + 1] - unified_col_boundaries[i]
                    for i in range(len(unified_col_boundaries) - 1)
                ]
            else:
                unified_col_boundaries = [0, 48000]
                unified_col_widths = [48000]

            if page:
                self.converter._apply_page_settings(ws, page)

            # 열 너비 적용
            for col_idx, width in enumerate(unified_col_widths, start=1):
                col_letter = get_column_letter(col_idx)
                excel_width = max(width / self.converter.HWPUNIT_TO_EXCEL_WIDTH, 1)
                ws.column_dimensions[col_letter].width = excel_width

            # 본문 배치
            current_row = 1
            for elem in body_elements:
                if elem["type"] == "para" and include_body:
                    cell = ws.cell(row=current_row, column=1, value=elem["text"])
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if len(unified_col_widths) > 1:
                        try:
                            ws.merge_cells(
                                start_row=current_row, start_column=1,
                                end_row=current_row, end_column=len(unified_col_widths)
                            )
                        except ValueError:
                            pass
                    current_row += 1

                elif elem["type"] == "table":
                    tbl_idx = elem["table_idx"]
                    if tbl_idx >= len(all_tables):
                        continue

                    table = all_tables[tbl_idx]
                    table_col_boundaries = self.converter._get_table_column_boundaries(table)
                    col_mapping = self.converter._map_columns_to_unified(table_col_boundaries, unified_col_boundaries)
                    cell_details = cell_details_map.get(tbl_idx, [])

                    # inline_nested 모드: nested 테이블을 부모 셀에 인라인 배치
                    nested_for_this = nested_by_parent.get(tbl_idx, []) if inline_nested else []

                    if nested_for_this:
                        # nested 테이블이 있으면 인라인 배치
                        rows_used, cell_mappings = self.converter.nested_handler.place_table_with_inline_nested(
                            ws, table, cell_details,
                            nested_for_this, all_tables, all_cell_details,
                            start_row=current_row,
                            parent_tbl_idx=tbl_idx,
                            get_column_widths_func=self.converter.placer.get_column_widths,
                            get_row_heights_func=self.converter.placer.get_row_heights,
                            apply_cell_style_func=self.converter.styler.apply_cell_style_single,
                            apply_merged_cell_borders_func=self.converter.styler.apply_merged_cell_borders,
                            hwp_color_to_rgb_func=self.converter.styler.hwp_color_to_rgb
                        )
                        # 시트 이름 정보 추가하여 매핑 저장
                        for mapping in cell_mappings:
                            mapping.sheet_name = sheet_name
                        all_cell_mappings.extend(cell_mappings)
                    elif split_by_para and cell_details:
                        rows_used = self.converter._place_table_with_para_split_unified(
                            ws, table, cell_details, col_mapping, unified_col_widths, current_row
                        )
                    else:
                        rows_used = self.converter._place_table_unified(
                            ws, table, cell_details, col_mapping, unified_col_widths, current_row
                        )

                    current_row += rows_used + 1

            # 페이지 설정
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        # 셀 정보 시트 추가 (옵션)
        if include_cell_info:
            try:
                from .cell_info_sheet import add_cell_info_sheet
            except ImportError:
                from cell_info_sheet import add_cell_info_sheet
            page_id = page.section_id if page and page.section_id else "section_0"
            add_cell_info_sheet(wb, hwpx_path, "CellInfo", hide_para_rows, page_id, "")

        # inline_nested 모드: 메타 시트 생성
        if inline_nested and all_cell_mappings:
            try:
                from .cell_info_sheet import add_meta_sheet_with_mappings
            except ImportError:
                from cell_info_sheet import add_meta_sheet_with_mappings
            add_meta_sheet_with_mappings(wb, all_cell_mappings, all_cell_details, "메타")

        # 기본 시트 삭제
        if default_sheet in wb:
            wb.remove(default_sheet)

        wb.save(output_path)
        print(f"  {len(wb.sheetnames)}개 시트 (북마크별) -> {output_path}")
        return output_path
