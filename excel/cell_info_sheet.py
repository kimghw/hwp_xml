# -*- coding: utf-8 -*-
"""
셀 정보 시트 생성 모듈

HWPX 셀 상세 정보를 Excel 시트로 저장합니다.
- list_id 기준 그룹핑
- para_id 행 숨김 처리
"""

import sys
from pathlib import Path
from typing import List, Dict, Optional, Union

# 프로젝트 루트 경로 설정
_project_root = Path(__file__).parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from hwpxml.get_cell_detail import GetCellDetail, CellDetail


# 컬럼 정의
CELL_INFO_COLUMNS = [
    ('page_id', 'page_id', 12),
    ('table_id', 'table_id', 12),
    ('list_id', 'list_id', 12),
    ('para_id', 'para_id', 12),
    ('row', 'row', 6),
    ('col', 'col', 6),
    ('end_row', 'end_row', 8),
    ('end_col', 'end_col', 8),
    ('row_span', 'rowspan', 8),
    ('col_span', 'colspan', 8),
    ('x', 'x', 8),
    ('y', 'y', 8),
    ('width', 'width', 10),
    ('height', 'height', 10),
    ('width_pt', 'width_pt', 10),
    ('height_pt', 'height_pt', 10),
    ('bg_color', 'bg_color', 10),
    ('margin_left', 'margin_l', 10),
    ('margin_right', 'margin_r', 10),
    ('margin_top', 'margin_t', 10),
    ('margin_bottom', 'margin_b', 10),
    ('font_name', 'font_name', 15),
    ('font_size_pt', 'font_size', 10),
    ('bold', 'bold', 6),
    ('italic', 'italic', 6),
    ('font_color', 'font_color', 12),
    ('underline', 'underline', 10),
    ('strikeout', 'strikeout', 10),
    ('align_h', 'align_h', 10),
    ('align_v', 'align_v', 10),
    ('line_spacing', 'line_spacing', 12),
    ('border_left', 'border_l', 10),
    ('border_right', 'border_r', 10),
    ('border_top', 'border_t', 10),
    ('border_bottom', 'border_b', 10),
    ('text', 'text', 30),
    ('field_name', 'field_name', 15),
    ('field_source', 'field_source', 15),
]


class CellInfoSheet:
    """
    셀 정보 시트 생성기

    사용 예:
        sheet_maker = CellInfoSheet()
        sheet_maker.add_to_workbook(wb, cells)
    """

    def __init__(self):
        self.cell_extractor = GetCellDetail()

    def create_cell_info_sheet(
        self,
        wb: Workbook,
        hwpx_path: Union[str, Path],
        sheet_name: str = "CellInfo",
        hide_para_rows: bool = True,
        page_id: str = "",
        table_id: str = ""
    ) -> Worksheet:
        """
        워크북에 셀 정보 시트 추가

        Args:
            wb: 대상 워크북
            hwpx_path: HWPX 파일 경로
            sheet_name: 시트 이름
            hide_para_rows: para_id 행 숨김 여부
            page_id: 페이지(섹션) ID
            table_id: 테이블 ID

        Returns:
            생성된 워크시트
        """
        # 셀 정보 추출
        cells = self.cell_extractor.from_hwpx(hwpx_path)

        # 시트 생성
        ws = wb.create_sheet(title=sheet_name)

        # 헤더 작성
        self._write_header(ws)

        # 데이터 작성 (list_id → para_id 계층 구조)
        self._write_data(ws, cells, hide_para_rows, page_id, table_id)

        return ws

    def _write_header(self, ws: Worksheet):
        """헤더 행 작성"""
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, (key, label, width) in enumerate(CELL_INFO_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

            # 열 너비 설정
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # 헤더 행 고정
        ws.freeze_panes = "A2"

    def _write_data(self, ws: Worksheet, cells: List[CellDetail], hide_para_rows: bool,
                    page_id: str = "", table_id: str = ""):
        """데이터 행 작성 (list_id → para_id 계층)"""
        # list_id별로 그룹핑
        grouped = self._group_by_list_id(cells)

        current_row = 2
        para_rows = []  # 숨길 행 번호들

        for list_id, cell_list in grouped.items():
            # 각 셀에 대해
            for cell in cell_list:
                # 1. 셀 정보 행 (메인)
                row_data = cell.to_dict()
                row_data['para_id'] = ''  # 셀 레벨에서는 para_id 비움
                row_data['page_id'] = page_id
                row_data['table_id'] = table_id

                self._write_row(ws, current_row, row_data, is_cell_row=True)
                current_row += 1

                # 2. 문단별 상세 정보 행
                for para in cell.paragraphs:
                    para_data = {
                        'page_id': '',  # 문단 레벨에서는 page_id 비움
                        'table_id': '',  # 문단 레벨에서는 table_id 비움
                        'list_id': '',  # 문단 레벨에서는 list_id 비움
                        'para_id': para.para_id,
                        'row': '',
                        'col': '',
                        'end_row': '',
                        'end_col': '',
                        'row_span': '',
                        'col_span': '',
                        'x': '',
                        'y': '',
                        'width': '',
                        'height': '',
                        'width_pt': '',
                        'height_pt': '',
                        'bg_color': '',
                        'margin_left': '',
                        'margin_right': '',
                        'margin_top': '',
                        'margin_bottom': '',
                        'font_name': cell.font.name,
                        'font_size_pt': cell.font.size_pt(),
                        'bold': cell.font.bold,
                        'italic': cell.font.italic,
                        'font_color': cell.font.color,
                        'underline': cell.font.underline,
                        'strikeout': cell.font.strikeout,
                        'align_h': para.align_h,
                        'align_v': para.align_v,
                        'line_spacing': para.line_spacing,
                        'border_left': '',
                        'border_right': '',
                        'border_top': '',
                        'border_bottom': '',
                        'text': '',
                        'field_name': '',
                        'field_source': '',
                    }

                    self._write_row(ws, current_row, para_data, is_cell_row=False)
                    para_rows.append(current_row)
                    current_row += 1

        # para_id 행 숨김 처리
        if hide_para_rows:
            for row_num in para_rows:
                ws.row_dimensions[row_num].hidden = True

    def _write_row(self, ws: Worksheet, row_num: int, data: Dict, is_cell_row: bool):
        """단일 행 작성"""
        # 스타일 설정
        if is_cell_row:
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        else:
            # para_id 행은 연한 회색
            fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        for col_idx, (key, label, width) in enumerate(CELL_INFO_COLUMNS, start=1):
            value = data.get(key, '')

            # bool 값 변환
            if isinstance(value, bool):
                value = 1 if value else 0

            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill = fill

            # 숫자 컬럼 오른쪽 정렬
            if key in ('row', 'col', 'end_row', 'end_col', 'row_span', 'col_span',
                       'x', 'y', 'width', 'height', 'width_pt', 'height_pt',
                       'margin_left', 'margin_right', 'margin_top', 'margin_bottom',
                       'font_size_pt', 'line_spacing', 'bold', 'italic', 'underline', 'strikeout'):
                cell.alignment = Alignment(horizontal="right")

    def _group_by_list_id(self, cells: List[CellDetail]) -> Dict[str, List[CellDetail]]:
        """list_id별로 셀 그룹핑"""
        grouped = {}
        for cell in cells:
            list_id = cell.list_id or 'unknown'
            if list_id not in grouped:
                grouped[list_id] = []
            grouped[list_id].append(cell)
        return grouped


def add_cell_info_sheet(
    wb: Workbook,
    hwpx_path: Union[str, Path],
    sheet_name: str = "CellInfo",
    hide_para_rows: bool = True,
    page_id: str = "",
    table_id: str = ""
) -> Worksheet:
    """
    워크북에 셀 정보 시트 추가 (편의 함수)

    Args:
        wb: 대상 워크북
        hwpx_path: HWPX 파일 경로
        sheet_name: 시트 이름
        hide_para_rows: para_id 행 숨김 여부
        page_id: 페이지(섹션) ID
        table_id: 테이블 ID

    Returns:
        생성된 워크시트
    """
    maker = CellInfoSheet()
    return maker.create_cell_info_sheet(wb, hwpx_path, sheet_name, hide_para_rows, page_id, table_id)


# 메타 시트 컬럼 정의 (Excel 위치 매핑 포함)
META_SHEET_COLUMNS = [
    ('sheet_name', 'sheet', 15),
    ('tbl_idx', 'tbl_idx', 8),
    ('table_id', 'table_id', 12),
    ('is_nested', 'is_nested', 10),
    ('parent_tbl_idx', 'parent_tbl', 10),
    ('orig_row', 'orig_row', 8),
    ('orig_col', 'orig_col', 8),
    ('excel_row', 'excel_row', 10),
    ('excel_col', 'excel_col', 10),
    ('list_id', 'list_id', 15),
    ('text', 'text', 30),
    ('field_name', 'field_name', 15),
]


def add_meta_sheet_with_mappings(
    wb: Workbook,
    cell_mappings: List,
    all_cell_details: List[List],
    sheet_name: str = "메타"
) -> Worksheet:
    """
    셀 위치 매핑 정보를 포함한 메타 시트 생성

    Args:
        wb: 대상 워크북
        cell_mappings: CellPositionMapping 리스트
        all_cell_details: 모든 테이블의 CellDetail 리스트
        sheet_name: 시트 이름

    Returns:
        생성된 워크시트
    """
    ws = wb.create_sheet(title=sheet_name)

    # 헤더 작성
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, (key, label, width) in enumerate(META_SHEET_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # 헤더 행 고정
    ws.freeze_panes = "A2"

    # 데이터 작성
    # list_id -> CellDetail 매핑 생성
    cell_detail_by_list_id = {}
    for tbl_details in all_cell_details:
        for cd in tbl_details:
            if cd.list_id:
                cell_detail_by_list_id[cd.list_id] = cd

    current_row = 2
    for mapping in cell_mappings:
        # CellDetail에서 추가 정보 가져오기
        cd = cell_detail_by_list_id.get(mapping.list_id)
        text = cd.text if cd else ""
        field_name = ""
        if cd and cd.paragraphs:
            for para in cd.paragraphs:
                if para.field_name:
                    field_name = para.field_name
                    break

        row_data = {
            'sheet_name': mapping.sheet_name if hasattr(mapping, 'sheet_name') else "",
            'tbl_idx': mapping.tbl_idx,
            'table_id': mapping.table_id,
            'is_nested': 1 if mapping.is_nested else 0,
            'parent_tbl_idx': mapping.parent_tbl_idx if mapping.parent_tbl_idx >= 0 else "",
            'orig_row': mapping.orig_row,
            'orig_col': mapping.orig_col,
            'excel_row': mapping.excel_row,
            'excel_col': mapping.excel_col,
            'list_id': mapping.list_id,
            'text': text,
            'field_name': field_name,
        }

        for col_idx, (key, label, width) in enumerate(META_SHEET_COLUMNS, start=1):
            value = row_data.get(key, '')
            ws.cell(row=current_row, column=col_idx, value=value)

        current_row += 1

    return ws


if __name__ == "__main__":
    import platform

    # 테스트
    if platform.system() == "Windows":
        hwpx_path = r"C:\hwp_xml\table.hwpx"
        output_path = r"C:\hwp_xml\excel\cell_info_test.xlsx"
    else:
        hwpx_path = "/mnt/c/hwp_xml/table.hwpx"
        output_path = "/mnt/c/hwp_xml/excel/cell_info_test.xlsx"

    print(f"입력: {hwpx_path}")
    print(f"출력: {output_path}")

    wb = Workbook()
    ws = add_cell_info_sheet(wb, hwpx_path)

    # 기본 시트 삭제
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    print(f"완료: {output_path}")
