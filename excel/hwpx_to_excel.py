# -*- coding: utf-8 -*-
"""HWPX 테이블 → Excel 변환 모듈"""

import sys
from pathlib import Path
from typing import Optional, Union, List, Dict

# 프로젝트 루트 경로 설정
_project_root = Path(__file__).parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.worksheet.properties import PageSetupProperties
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass

from hwpxml.get_table_property import GetTableProperty, TableProperty
from hwpxml.get_page_property import GetPageProperty, PageProperty, Unit
from hwpxml.get_cell_detail import GetCellDetail, CellDetail
from excel.styles import ExcelStyler
from excel.table_placement import TablePlacer
from excel.bookmark import BookmarkHandler
from excel.nested_table import NestedTableHandler, TableHierarchy


class HwpxToExcel:
    """
    HWPX 테이블을 Excel로 변환

    사용 예:
        converter = HwpxToExcel()
        converter.convert("input.hwpx", "output.xlsx")

        # 북마크 기반 추출
        bookmarks = converter.get_bookmarks("input.hwpx")
        converter.convert_by_bookmark("input.hwpx", "7. 시장 사업화 계획")
    """

    # 엑셀 열 너비 변환 계수
    # 엑셀 열 너비 1 문자 ≈ 8.43 픽셀 (Calibri 11pt 기준)
    # 1 HWPUNIT = 1/100 pt, 1 pt ≈ 1.333 픽셀 (96 DPI)
    # HWPUNIT → 픽셀: hwpunit / 100 * 1.333 = hwpunit / 75
    # 픽셀 → 엑셀 문자: pixel / 8.43
    # 따라서 HWPUNIT → 엑셀 열 너비: hwpunit / 75 / 8.43 ≈ hwpunit / 632
    HWPUNIT_TO_EXCEL_WIDTH = 632

    # 엑셀 행 높이는 포인트 단위
    # HWPUNIT → pt: hwpunit / 100
    HWPUNIT_TO_PT = 100

    def __init__(self):
        self.table_parser = GetTableProperty()
        self.page_parser = GetPageProperty()
        self.cell_detail_parser = GetCellDetail()
        self.table_hierarchy: List[TableHierarchy] = []
        self.styler = ExcelStyler()
        self.placer = TablePlacer()
        self.nested_handler = NestedTableHandler()
        self.bookmark_handler = BookmarkHandler(self)

    def get_bookmarks(self, hwpx_path: Union[str, Path]) -> List[dict]:
        """HWPX 파일에서 북마크 목록 추출 (BookmarkHandler로 위임)"""
        return self.bookmark_handler.get_bookmarks(hwpx_path)

    def get_bookmark_table_mapping(self, hwpx_path: Union[str, Path]) -> Dict[str, List[int]]:
        """북마크별로 소속 테이블 인덱스 매핑 (BookmarkHandler로 위임)"""
        return self.bookmark_handler.get_bookmark_table_mapping(hwpx_path)

    def convert_by_bookmark(
        self,
        hwpx_path: Union[str, Path],
        bookmark_name: str,
        output_path: Optional[Union[str, Path]] = None,
        split_by_para: bool = False,
        include_cell_info: bool = False,
        hide_para_rows: bool = True
    ) -> Path:
        """
        특정 북마크 섹션의 테이블만 하나의 시트로 변환

        Args:
            hwpx_path: HWPX 파일 경로
            bookmark_name: 북마크 이름 (부분 일치 지원)
            output_path: 출력 Excel 경로
            split_by_para: 문단별 행 분할 여부
            include_cell_info: 셀 상세 정보 시트 포함 여부
            hide_para_rows: para_id 행 숨김 여부 (include_cell_info=True일 때)

        Returns:
            생성된 Excel 파일 경로
        """
        hwpx_path = Path(hwpx_path)

        # 북마크-테이블 매핑 가져오기
        bookmark_mapping = self.get_bookmark_table_mapping(hwpx_path)

        # 북마크 이름 찾기 (부분 일치)
        matched_bookmark = None
        for bm_name in bookmark_mapping.keys():
            if bookmark_name in bm_name:
                matched_bookmark = bm_name
                break

        if not matched_bookmark:
            raise ValueError(f"북마크를 찾을 수 없습니다: {bookmark_name}")

        table_indices = bookmark_mapping[matched_bookmark]
        if not table_indices:
            raise ValueError(f"북마크 '{matched_bookmark}'에 테이블이 없습니다")

        # 출력 경로 설정
        if output_path is None:
            safe_name = bookmark_name.replace(" ", "_").replace("/", "_")[:20]
            output_path = hwpx_path.with_name(f"{hwpx_path.stem}_{safe_name}.xlsx")
        else:
            output_path = Path(output_path)

        # 모든 테이블 데이터 로드
        all_tables = self.table_parser.from_hwpx(hwpx_path)
        pages = self.page_parser.from_hwpx(hwpx_path)
        all_cell_details = self.cell_detail_parser.from_hwpx_by_table(hwpx_path)

        # 해당 북마크의 테이블만 필터링
        tables = [all_tables[i] for i in table_indices if i < len(all_tables)]
        cell_details_list = [all_cell_details[i] for i in table_indices if i < len(all_cell_details)]

        if not tables:
            raise ValueError(f"테이블 인덱스가 범위를 벗어났습니다: {table_indices}")

        page = pages[0] if pages else None

        # 통합 열 그리드 생성
        unified_col_boundaries = self.placer.build_unified_column_grid(tables)
        unified_col_widths = []
        for i in range(len(unified_col_boundaries) - 1):
            unified_col_widths.append(unified_col_boundaries[i + 1] - unified_col_boundaries[i])

        # Excel 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = matched_bookmark[:31] if len(matched_bookmark) <= 31 else matched_bookmark[:28] + "..."

        if page:
            self.placer.apply_page_settings(ws, page)

        # 통합 열 너비 적용
        for col_idx, width in enumerate(unified_col_widths, start=1):
            col_letter = get_column_letter(col_idx)
            excel_width = width / self.placer.HWPUNIT_TO_EXCEL_WIDTH
            excel_width = max(excel_width, 1)
            ws.column_dimensions[col_letter].width = excel_width

        # 각 테이블 배치
        current_row = 1
        for tbl_idx, table in enumerate(tables):
            table_col_boundaries = self.placer.get_table_column_boundaries(table)
            col_mapping = self.placer.map_columns_to_unified(table_col_boundaries, unified_col_boundaries)
            cell_details = cell_details_list[tbl_idx] if tbl_idx < len(cell_details_list) else []

            if split_by_para and cell_details:
                rows_used = self.placer.place_table_with_para_split_unified(
                    ws, table, cell_details, col_mapping, unified_col_widths, current_row
                )
            else:
                rows_used = self.placer.place_table_unified(
                    ws, table, cell_details, col_mapping, unified_col_widths, current_row
                )

            current_row += rows_used + 1  # 테이블 사이 빈 행

        # 모든 열을 1페이지 폭에 맞추기
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 높이는 제한 없음
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        # 셀 정보 시트 추가 (옵션)
        if include_cell_info:
            try:
                from .cell_info_sheet import add_cell_info_sheet
            except ImportError:
                from cell_info_sheet import add_cell_info_sheet
            page_id = page.section_id if page and page.section_id else "section_0"
            add_cell_info_sheet(wb, hwpx_path, f"{matched_bookmark[:20]}_메타", hide_para_rows, page_id, "")

        wb.save(output_path)
        print(f"  북마크 '{matched_bookmark}': {len(tables)}개 테이블 → {output_path}")
        return output_path

    def get_body_elements(self, hwpx_path: Union[str, Path]) -> List[dict]:
        """HWPX 파일에서 본문 요소(문단, 테이블) 순서대로 추출 (BookmarkHandler로 위임)"""
        return self.bookmark_handler.get_body_elements(hwpx_path)

    def get_bookmark_body_mapping(self, hwpx_path: Union[str, Path]) -> Dict[str, List[dict]]:
        """북마크별로 본문 요소(문단, 테이블) 매핑 (BookmarkHandler로 위임)"""
        return self.bookmark_handler.get_bookmark_body_mapping(hwpx_path)

    def convert_by_bookmark_with_body(
        self,
        hwpx_path: Union[str, Path],
        bookmark_name: str,
        output_path: Optional[Union[str, Path]] = None,
        split_by_para: bool = False,
        include_cell_info: bool = False,
        hide_para_rows: bool = True
    ) -> Path:
        """
        특정 북마크 섹션의 본문(문단+테이블)을 하나의 시트로 변환

        본문 문단은 첫 열에 배치하고, 테이블은 통합 열 그리드에 맞춰 배치합니다.

        Args:
            hwpx_path: HWPX 파일 경로
            bookmark_name: 북마크 이름 (부분 일치 지원)
            output_path: 출력 Excel 경로
            split_by_para: 문단별 행 분할 여부
            include_cell_info: 셀 상세 정보 시트 포함 여부
            hide_para_rows: para_id 행 숨김 여부 (include_cell_info=True일 때)

        Returns:
            생성된 Excel 파일 경로
        """
        hwpx_path = Path(hwpx_path)

        # 북마크-본문 매핑 가져오기
        bookmark_body_mapping = self.get_bookmark_body_mapping(hwpx_path)

        # 북마크 이름 찾기 (부분 일치)
        matched_bookmark = None
        for bm_name in bookmark_body_mapping.keys():
            if bookmark_name in bm_name:
                matched_bookmark = bm_name
                break

        if not matched_bookmark:
            raise ValueError(f"북마크를 찾을 수 없습니다: {bookmark_name}")

        body_elements = bookmark_body_mapping[matched_bookmark]
        if not body_elements:
            raise ValueError(f"북마크 '{matched_bookmark}'에 내용이 없습니다")

        # 출력 경로 설정
        if output_path is None:
            safe_name = bookmark_name.replace(" ", "_").replace("/", "_")[:20]
            output_path = hwpx_path.with_name(f"{hwpx_path.stem}_{safe_name}_body.xlsx")
        else:
            output_path = Path(output_path)

        # 테이블 인덱스 추출
        table_indices = [e["table_idx"] for e in body_elements if e["type"] == "table"]

        # 모든 테이블 데이터 로드
        all_tables = self.table_parser.from_hwpx(hwpx_path)
        pages = self.page_parser.from_hwpx(hwpx_path)
        all_cell_details = self.cell_detail_parser.from_hwpx_by_table(hwpx_path)

        # 해당 북마크의 테이블만 필터링
        tables = [all_tables[i] for i in table_indices if i < len(all_tables)]
        cell_details_map = {i: all_cell_details[i] for i in table_indices if i < len(all_cell_details)}

        page = pages[0] if pages else None

        # 통합 열 그리드 생성 (테이블이 있을 경우)
        if tables:
            unified_col_boundaries = self.placer.build_unified_column_grid(tables)
            unified_col_widths = []
            for i in range(len(unified_col_boundaries) - 1):
                unified_col_widths.append(unified_col_boundaries[i + 1] - unified_col_boundaries[i])
        else:
            # 테이블이 없으면 기본 열 1개
            unified_col_boundaries = [0, 48000]  # A4 폭 정도
            unified_col_widths = [48000]

        # Excel 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = matched_bookmark[:31] if len(matched_bookmark) <= 31 else matched_bookmark[:28] + "..."

        if page:
            self.placer.apply_page_settings(ws, page)

        # 통합 열 너비 적용
        for col_idx, width in enumerate(unified_col_widths, start=1):
            col_letter = get_column_letter(col_idx)
            excel_width = width / self.placer.HWPUNIT_TO_EXCEL_WIDTH
            excel_width = max(excel_width, 1)
            ws.column_dimensions[col_letter].width = excel_width

        # 본문 요소 순서대로 배치
        current_row = 1
        for elem in body_elements:
            if elem["type"] == "para":
                # 본문 문단: 첫 열에 텍스트 배치, 전체 열 병합
                cell = ws.cell(row=current_row, column=1, value=elem["text"])
                cell.alignment = Alignment(wrap_text=True, vertical='top')

                # 전체 열 병합
                if len(unified_col_widths) > 1:
                    try:
                        ws.merge_cells(
                            start_row=current_row,
                            start_column=1,
                            end_row=current_row,
                            end_column=len(unified_col_widths)
                        )
                    except ValueError:
                        pass

                current_row += 1

            elif elem["type"] == "table":
                tbl_idx = elem["table_idx"]
                if tbl_idx >= len(all_tables):
                    continue

                table = all_tables[tbl_idx]
                table_col_boundaries = self.placer.get_table_column_boundaries(table)
                col_mapping = self.placer.map_columns_to_unified(table_col_boundaries, unified_col_boundaries)
                cell_details = cell_details_map.get(tbl_idx, [])

                if split_by_para and cell_details:
                    rows_used = self.placer.place_table_with_para_split_unified(
                        ws, table, cell_details, col_mapping, unified_col_widths, current_row
                    )
                else:
                    rows_used = self.placer.place_table_unified(
                        ws, table, cell_details, col_mapping, unified_col_widths, current_row
                    )

                current_row += rows_used + 1  # 테이블 후 빈 행

        # 모든 열을 1페이지 폭에 맞추기
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        # 셀 정보 시트 추가 (옵션)
        if include_cell_info:
            try:
                from .cell_info_sheet import add_cell_info_sheet
            except ImportError:
                from cell_info_sheet import add_cell_info_sheet
            page_id = page.section_id if page and page.section_id else "section_0"
            add_cell_info_sheet(wb, hwpx_path, f"{matched_bookmark[:20]}_메타", hide_para_rows, page_id, "")

        wb.save(output_path)
        para_count = sum(1 for e in body_elements if e["type"] == "para")
        table_count = sum(1 for e in body_elements if e["type"] == "table")
        print(f"  북마크 '{matched_bookmark}': 본문 {para_count}개 + 테이블 {table_count}개 → {output_path}")
        return output_path

    def convert_all_by_bookmark(
        self,
        hwpx_path: Union[str, Path],
        output_path: Optional[Union[str, Path]] = None,
        include_body: bool = True,
        split_by_para: bool = False,
        include_cell_info: bool = False,
        hide_para_rows: bool = True,
        inline_nested: bool = False
    ) -> Path:
        """
        전체 문서를 북마크별로 시트 분리하여 변환

        각 북마크 섹션이 별도의 시트로 생성됩니다.

        Args:
            hwpx_path: HWPX 파일 경로
            output_path: 출력 Excel 경로
            include_body: 본문 문단 포함 여부 (False면 테이블만)
            split_by_para: 문단별 행 분할 여부
            include_cell_info: 셀 상세 정보 시트 포함 여부
            hide_para_rows: para_id 행 숨김 여부 (include_cell_info=True일 때)
            inline_nested: True면 nested 테이블을 부모 셀에 인라인 배치

        Returns:
            생성된 Excel 파일 경로
        """
        hwpx_path = Path(hwpx_path)

        if output_path is None:
            suffix = "_by_bookmark_body" if include_body else "_by_bookmark"
            if inline_nested:
                suffix += "_inline"
            output_path = hwpx_path.with_name(f"{hwpx_path.stem}{suffix}.xlsx")
        else:
            output_path = Path(output_path)

        # 데이터 로드
        bookmark_body_mapping = self.get_bookmark_body_mapping(hwpx_path)
        all_tables = self.table_parser.from_hwpx(hwpx_path)
        pages = self.page_parser.from_hwpx(hwpx_path)
        all_cell_details = self.cell_detail_parser.from_hwpx_by_table(hwpx_path)
        page = pages[0] if pages else None

        # inline_nested 모드: 테이블 계층 구조 파악
        table_hierarchy = []
        nested_by_parent = {}  # parent_tbl_idx -> [nested1, nested2, ...]
        if inline_nested:
            table_hierarchy = self._parse_table_hierarchy(hwpx_path)
            for h in table_hierarchy:
                if h.parent_tbl_idx != -1:
                    if h.parent_tbl_idx not in nested_by_parent:
                        nested_by_parent[h.parent_tbl_idx] = []
                    nested_by_parent[h.parent_tbl_idx].append(h)

        wb = Workbook()
        default_sheet = wb.active

        for bm_idx, (bm_name, body_elements) in enumerate(bookmark_body_mapping.items()):
            if not body_elements:
                continue

            # 시트 이름 (31자 제한, 유효하지 않은 문자 제거, 중복 방지)
            # Excel 시트 이름에 사용할 수 없는 문자: [ ] * ? / \ :
            safe_name = bm_name
            for invalid_char in ['[', ']', '*', '?', '/', '\\', ':']:
                safe_name = safe_name.replace(invalid_char, '_')
            sheet_name = safe_name[:31] if len(safe_name) <= 31 else safe_name[:28] + "..."
            if sheet_name in wb.sheetnames:
                sheet_name = f"{bm_idx}_{sheet_name}"[:31]

            ws = wb.create_sheet(title=sheet_name)

            # 테이블 인덱스 추출
            table_indices = [e["table_idx"] for e in body_elements if e["type"] == "table"]
            tables = [all_tables[i] for i in table_indices if i < len(all_tables)]
            cell_details_map = {i: all_cell_details[i] for i in table_indices if i < len(all_cell_details)}

            # 통합 열 그리드 생성
            if tables:
                unified_col_boundaries = self.placer.build_unified_column_grid(tables)
                unified_col_widths = [
                    unified_col_boundaries[i + 1] - unified_col_boundaries[i]
                    for i in range(len(unified_col_boundaries) - 1)
                ]
            else:
                unified_col_boundaries = [0, 48000]
                unified_col_widths = [48000]

            if page:
                self.placer.apply_page_settings(ws, page)

            # 열 너비 적용
            for col_idx, width in enumerate(unified_col_widths, start=1):
                col_letter = get_column_letter(col_idx)
                excel_width = max(width / self.placer.HWPUNIT_TO_EXCEL_WIDTH, 1)
                ws.column_dimensions[col_letter].width = excel_width

            # 본문 배치
            current_row = 1
            for elem in body_elements:
                if elem["type"] == "para" and include_body:
                    cell = ws.cell(row=current_row, column=1, value=elem["text"])
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if len(unified_col_widths) > 1:
                        try:
                            ws.merge_cells(
                                start_row=current_row, start_column=1,
                                end_row=current_row, end_column=len(unified_col_widths)
                            )
                        except ValueError:
                            pass
                    current_row += 1

                elif elem["type"] == "table":
                    tbl_idx = elem["table_idx"]
                    if tbl_idx >= len(all_tables):
                        continue

                    table = all_tables[tbl_idx]
                    table_col_boundaries = self.placer.get_table_column_boundaries(table)
                    col_mapping = self.placer.map_columns_to_unified(table_col_boundaries, unified_col_boundaries)
                    cell_details = cell_details_map.get(tbl_idx, [])

                    # inline_nested 모드: nested 테이블을 부모 셀에 인라인 배치
                    nested_for_this = nested_by_parent.get(tbl_idx, []) if inline_nested else []

                    if nested_for_this:
                        # nested 테이블이 있으면 인라인 배치
                        rows_used = self.nested_handler.place_table_with_inline_nested(
                            ws, table, cell_details,
                            nested_for_this, all_tables, all_cell_details,
                            start_row=current_row,
                            get_column_widths_func=self.placer.get_column_widths,
                            get_row_heights_func=self.placer.get_row_heights,
                            apply_cell_style_func=self.styler.apply_cell_style_single,
                            apply_merged_cell_borders_func=self.styler.apply_merged_cell_borders,
                            hwp_color_to_rgb_func=self.styler.hwp_color_to_rgb
                        )
                    elif split_by_para and cell_details:
                        rows_used = self.placer.place_table_with_para_split_unified(
                            ws, table, cell_details, col_mapping, unified_col_widths, current_row
                        )
                    else:
                        rows_used = self.placer.place_table_unified(
                            ws, table, cell_details, col_mapping, unified_col_widths, current_row
                        )

                    current_row += rows_used + 1

            # 페이지 설정
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        # 셀 정보 시트 추가 (옵션)
        if include_cell_info:
            try:
                from .cell_info_sheet import add_cell_info_sheet
            except ImportError:
                from cell_info_sheet import add_cell_info_sheet
            page_id = page.section_id if page and page.section_id else "section_0"
            add_cell_info_sheet(wb, hwpx_path, "CellInfo", hide_para_rows, page_id, "")

        # 기본 시트 삭제
        if default_sheet in wb:
            wb.remove(default_sheet)

        wb.save(output_path)
        print(f"  {len(wb.sheetnames)}개 시트 (북마크별) → {output_path}")
        return output_path

    def _parse_table_hierarchy(self, hwpx_path: Union[str, Path]) -> List[TableHierarchy]:
        """HWPX에서 테이블 계층 구조 파악 (NestedTableHandler로 위임)"""
        return self.nested_handler.parse_table_hierarchy(hwpx_path)

    def convert(
        self,
        hwpx_path: Union[str, Path],
        output_path: Optional[Union[str, Path]] = None,
        table_index: int = 0,
        include_cell_info: bool = False,
        hide_para_rows: bool = True
    ) -> Path:
        """
        HWPX 파일의 테이블을 Excel로 변환

        Args:
            hwpx_path: HWPX 파일 경로
            output_path: 출력 Excel 경로 (없으면 자동 생성)
            table_index: 변환할 테이블 인덱스 (기본 0)
            include_cell_info: 셀 상세 정보 시트 포함 여부
            hide_para_rows: para_id 행 숨김 여부 (include_cell_info=True일 때)

        Returns:
            생성된 Excel 파일 경로
        """
        hwpx_path = Path(hwpx_path)

        if output_path is None:
            output_path = hwpx_path.with_suffix('.xlsx')
        else:
            output_path = Path(output_path)

        # HWPX에서 데이터 추출
        tables = self.table_parser.from_hwpx(hwpx_path)
        pages = self.page_parser.from_hwpx(hwpx_path)

        if not tables:
            raise ValueError(f"테이블을 찾을 수 없습니다: {hwpx_path}")

        table = tables[table_index] if table_index < len(tables) else tables[0]
        page = pages[0] if pages else None

        # Excel 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "본문"

        # 1. 페이지 설정 적용
        if page:
            self.placer.apply_page_settings(ws, page)

        # 2. 열 너비 설정
        self.placer.apply_column_widths(ws, table)

        # 3. 행 높이 설정
        self.placer.apply_row_heights(ws, table)

        # 4. 셀 병합 처리
        self.placer.apply_cell_merges(ws, table)

        # 5. 셀 정보 시트 추가 (옵션)
        if include_cell_info:
            try:
                from .cell_info_sheet import add_cell_info_sheet
            except ImportError:
                from cell_info_sheet import add_cell_info_sheet
            # 페이지, 테이블 ID 추출
            page_id = page.section_id if page and page.section_id else "section_0"
            table_id = str(table.id) if table and table.id else f"table_{table_index}"
            add_cell_info_sheet(wb, hwpx_path, "본문_메타", hide_para_rows, page_id, table_id)

        # 저장
        wb.save(output_path)
        return output_path

    def convert_all_to_single_sheet(
        self,
        hwpx_path: Union[str, Path],
        output_path: Optional[Union[str, Path]] = None,
        split_by_para: bool = False
    ) -> Path:
        """
        HWPX 파일의 모든 테이블을 하나의 시트에 통합 변환

        각 테이블의 열 너비가 다르므로, 모든 열 경계를 수집하여
        통합 열 그리드를 생성하고 셀을 매핑합니다.
        테이블의 원본 너비를 유지하며 스케일링하지 않습니다.

        Args:
            hwpx_path: HWPX 파일 경로
            output_path: 출력 Excel 경로 (없으면 자동 생성)
            split_by_para: 셀 내 문단별로 행 분할 여부

        Returns:
            생성된 Excel 파일 경로
        """
        hwpx_path = Path(hwpx_path)

        if output_path is None:
            output_path = hwpx_path.with_name(hwpx_path.stem + "_all.xlsx")
        else:
            output_path = Path(output_path)

        # HWPX에서 데이터 추출
        tables = self.table_parser.from_hwpx(hwpx_path)
        pages = self.page_parser.from_hwpx(hwpx_path)
        table_cell_details = self.cell_detail_parser.from_hwpx_by_table(hwpx_path)

        if not tables:
            raise ValueError(f"테이블을 찾을 수 없습니다: {hwpx_path}")

        page = pages[0] if pages else None

        # 1. 모든 테이블의 열 경계 수집 및 통합 그리드 생성 (원본 너비 유지)
        unified_col_boundaries = self.placer.build_unified_column_grid(tables)
        # 통합 열 너비 계산
        unified_col_widths = []
        for i in range(len(unified_col_boundaries) - 1):
            unified_col_widths.append(unified_col_boundaries[i + 1] - unified_col_boundaries[i])

        # Excel 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "통합"

        # 페이지 설정
        if page:
            self.placer.apply_page_settings(ws, page)

        # 통합 열 너비 적용
        for col_idx, width in enumerate(unified_col_widths, start=1):
            col_letter = get_column_letter(col_idx)
            excel_width = width / self.placer.HWPUNIT_TO_EXCEL_WIDTH
            excel_width = max(excel_width, 1)
            ws.column_dimensions[col_letter].width = excel_width

        # 2. 각 테이블을 순서대로 배치
        current_row = 1
        for tbl_idx, table in enumerate(tables):
            # 이 테이블의 열 경계 계산 (원본 너비)
            table_col_boundaries = self.placer.get_table_column_boundaries(table)

            # 테이블 원본 열 → 통합 열 매핑
            col_mapping = self.placer.map_columns_to_unified(table_col_boundaries, unified_col_boundaries)

            # 셀 디테일
            cell_details = table_cell_details[tbl_idx] if tbl_idx < len(table_cell_details) else []

            # split_by_para에 따른 처리
            if split_by_para and cell_details:
                rows_used = self.placer.place_table_with_para_split_unified(
                    ws, table, cell_details, col_mapping, unified_col_widths, current_row
                )
            else:
                rows_used = self.placer.place_table_unified(
                    ws, table, cell_details, col_mapping, unified_col_widths, current_row
                )

            current_row += rows_used + 1  # 테이블 사이 빈 행 추가

        # 저장
        wb.save(output_path)
        print(f"  {len(tables)}개 테이블 → 1개 시트 (통합 열: {len(unified_col_widths)}개)")
        return output_path

    def convert_all(
        self,
        hwpx_path: Union[str, Path],
        output_path: Optional[Union[str, Path]] = None,
        include_cell_info: bool = False,
        hide_para_rows: bool = True,
        split_by_para: bool = False,
        inline_nested: bool = False
    ) -> Path:
        """
        HWPX 파일의 모든 테이블을 Excel 시트로 변환
        - 최상위 테이블: 개별 시트로 생성
        - 중첩 테이블:
          - inline_nested=False (기본): tbl_sub 시트에 정리 + 부모 셀에 하이퍼링크/메모
          - inline_nested=True: 부모 셀에 인라인으로 표시 (경계선으로 구분)

        Args:
            hwpx_path: HWPX 파일 경로
            output_path: 출력 Excel 경로 (없으면 자동 생성)
            include_cell_info: 셀 상세 정보 시트 포함 여부
            hide_para_rows: para_id 행 숨김 여부
            split_by_para: 셀 내 문단별로 행 분할 (True면 문단 수만큼 행 생성)
            inline_nested: 중첩 테이블을 부모 셀에 인라인으로 표시 (True면 경계선으로 구분)

        Returns:
            생성된 Excel 파일 경로
        """
        hwpx_path = Path(hwpx_path)

        if output_path is None:
            output_path = hwpx_path.with_suffix('.xlsx')
        else:
            output_path = Path(output_path)

        # HWPX에서 데이터 추출
        tables = self.table_parser.from_hwpx(hwpx_path)
        pages = self.page_parser.from_hwpx(hwpx_path)
        # 테이블별로 그룹화된 셀 디테일 가져오기
        table_cell_details = self.cell_detail_parser.from_hwpx_by_table(hwpx_path)

        if not tables:
            raise ValueError(f"테이블을 찾을 수 없습니다: {hwpx_path}")

        # 테이블 계층 구조 파악
        self.table_hierarchy = self._parse_table_hierarchy(hwpx_path)

        # 최상위 테이블과 중첩 테이블 분리
        top_level_tables = [h for h in self.table_hierarchy if h.parent_tbl_idx == -1]
        nested_tables = [h for h in self.table_hierarchy if h.parent_tbl_idx != -1]

        page = pages[0] if pages else None

        # Excel 워크북 생성
        wb = Workbook()
        default_sheet = wb.active

        # inline_nested 모드 처리
        if inline_nested:
            # === inline_nested=True: 최상위 테이블만 시트 생성, nested는 부모 셀에 인라인 배치 ===

            # 부모 테이블별 nested 테이블 그룹화
            nested_by_parent = {}  # parent_tbl_idx -> [nested1, nested2, ...]
            for nested in nested_tables:
                parent_idx = nested.parent_tbl_idx
                if parent_idx not in nested_by_parent:
                    nested_by_parent[parent_idx] = []
                nested_by_parent[parent_idx].append(nested)

            # 최상위 테이블만 시트 생성
            for top_h in top_level_tables:
                tbl_idx = top_h.tbl_idx
                if tbl_idx >= len(tables):
                    continue

                table = tables[tbl_idx]
                cell_details = table_cell_details[tbl_idx] if tbl_idx < len(table_cell_details) else []

                ws = wb.create_sheet(title=f"tbl_{tbl_idx}")

                if page:
                    self.placer.apply_page_settings(ws, page)

                # 이 테이블에 속한 nested 테이블 목록
                nested_for_this = nested_by_parent.get(tbl_idx, [])

                if nested_for_this:
                    # nested 테이블이 있으면 인라인 배치 함수 사용
                    self.nested_handler.place_table_with_inline_nested(
                        ws, table, cell_details,
                        nested_for_this, tables, table_cell_details,
                        start_row=1,
                        get_column_widths_func=self.placer.get_column_widths,
                        get_row_heights_func=self.placer.get_row_heights,
                        apply_cell_style_func=self.styler.apply_cell_style_single,
                        apply_merged_cell_borders_func=self.styler.apply_merged_cell_borders,
                        hwp_color_to_rgb_func=self.styler.hwp_color_to_rgb
                    )
                else:
                    # nested 테이블이 없으면 기존 방식
                    self.placer.apply_column_widths(ws, table)
                    if split_by_para and cell_details:
                        self.placer.apply_table_with_para_split(ws, table, cell_details)
                    else:
                        self.placer.apply_row_heights(ws, table)
                        self.placer.apply_cell_merges(ws, table)
                        self.styler.apply_cell_styles(ws, cell_details, tbl_idx)

                # 페이지 설정
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        else:
            # === inline_nested=False (기본): 모든 테이블 별도 시트 + 하이퍼링크 ===

            # 1. 모든 테이블 시트 생성 (최상위 + 중첩 모두)
            # 각 테이블의 row_offset_map 저장 (split_by_para용)
            table_row_offset_maps = {}

            for idx, table in enumerate(tables):
                ws = wb.create_sheet(title=f"tbl_{idx}")

                if page:
                    self.placer.apply_page_settings(ws, page)

                self.placer.apply_column_widths(ws, table)

                # split_by_para 옵션에 따라 처리
                if split_by_para and idx < len(table_cell_details):
                    # 문단별 행 분할 (row_offset_map 반환)
                    row_offset_map = self.placer.apply_table_with_para_split(ws, table, table_cell_details[idx])
                    table_row_offset_maps[idx] = row_offset_map
                else:
                    # 기존 방식
                    self.placer.apply_row_heights(ws, table)
                    self.placer.apply_cell_merges(ws, table)

                    # 셀 스타일 적용
                    if idx < len(table_cell_details):
                        self.styler.apply_cell_styles(ws, table_cell_details[idx], idx)

            # 2. 부모 테이블 셀에 중첩 테이블 하이퍼링크/메모 추가
            # 같은 셀에 여러 nested 테이블이 있을 수 있으므로 그룹화
            nested_by_cell = {}  # (parent_idx, row, col) -> [nested1, nested2, ...]
            for nested in nested_tables:
                parent_idx = nested.parent_tbl_idx
                if parent_idx < 0 or parent_idx >= len(tables):
                    continue

                orig_row = nested.parent_row
                col = nested.parent_col + 1  # 1-based

                # split_by_para일 때 row_offset_map 사용
                if parent_idx in table_row_offset_maps:
                    row_offset_map = table_row_offset_maps[parent_idx]
                    row = row_offset_map.get(orig_row, orig_row) + 1  # 1-based
                else:
                    row = orig_row + 1  # 1-based

                key = (parent_idx, row, col)
                if key not in nested_by_cell:
                    nested_by_cell[key] = []
                nested_by_cell[key].append(nested)

            # 그룹화된 nested 테이블 처리
            for (parent_idx, row, col), nested_list in nested_by_cell.items():
                parent_sheet_name = f"tbl_{parent_idx}"
                if parent_sheet_name not in wb.sheetnames:
                    continue

                parent_ws = wb[parent_sheet_name]

                if row > 0 and col > 0:
                    cell = parent_ws.cell(row=row, column=col)

                    # 여러 nested 테이블이 있을 경우 표시
                    if len(nested_list) == 1:
                        nested = nested_list[0]
                        cell.hyperlink = f"#tbl_{nested.tbl_idx}!A1"
                        cell.value = f"→tbl_{nested.tbl_idx}"
                        comment_text = (
                            f"중첩 테이블: tbl_{nested.tbl_idx}\n"
                            f"크기: {nested.row_count}행 x {nested.col_count}열"
                        )
                    else:
                        # 여러 개: 첫 번째 테이블로 링크, 나머지는 메모에 표시
                        tbl_names = [f"tbl_{n.tbl_idx}" for n in nested_list]
                        cell.hyperlink = f"#tbl_{nested_list[0].tbl_idx}!A1"
                        cell.value = f"→{','.join(tbl_names)}"
                        comment_lines = []
                        for n in nested_list:
                            comment_lines.append(f"tbl_{n.tbl_idx}: {n.row_count}행 x {n.col_count}열")
                        comment_text = "중첩 테이블:\n" + "\n".join(comment_lines)

                    cell.style = "Hyperlink"
                    cell.comment = Comment(comment_text, "System")

            # 3. tbl_sub 시트 생성 (중첩 테이블이 있는 경우)
            if nested_tables:
                ws_sub = wb.create_sheet(title="tbl_sub")

                # 헤더
                headers = ["tbl_idx", "table_id", "parent_tbl", "parent_row", "parent_col", "rows", "cols", "link"]
                for col_idx, header in enumerate(headers, 1):
                    ws_sub.cell(row=1, column=col_idx, value=header)

                # 데이터
                for row_idx, nested in enumerate(nested_tables, 2):
                    ws_sub.cell(row=row_idx, column=1, value=nested.tbl_idx)
                    ws_sub.cell(row=row_idx, column=2, value=nested.table_id)
                    ws_sub.cell(row=row_idx, column=3, value=f"tbl_{nested.parent_tbl_idx}")
                    ws_sub.cell(row=row_idx, column=4, value=nested.parent_row)
                    ws_sub.cell(row=row_idx, column=5, value=nested.parent_col)
                    ws_sub.cell(row=row_idx, column=6, value=nested.row_count)
                    ws_sub.cell(row=row_idx, column=7, value=nested.col_count)

                    # 링크 셀
                    link_cell = ws_sub.cell(row=row_idx, column=8)
                    link_cell.hyperlink = f"#tbl_{nested.tbl_idx}!A1"
                    link_cell.value = f"→tbl_{nested.tbl_idx}"
                    link_cell.style = "Hyperlink"

                # 열 너비 조정
                ws_sub.column_dimensions['A'].width = 8
                ws_sub.column_dimensions['B'].width = 15
                ws_sub.column_dimensions['C'].width = 12
                ws_sub.column_dimensions['D'].width = 10
                ws_sub.column_dimensions['E'].width = 10
                ws_sub.column_dimensions['F'].width = 6
                ws_sub.column_dimensions['G'].width = 6
                ws_sub.column_dimensions['H'].width = 12

        # 기본 시트 삭제
        wb.remove(default_sheet)

        # 셀 정보 시트 추가 (옵션)
        if include_cell_info:
            try:
                from .cell_info_sheet import add_cell_info_sheet
            except ImportError:
                from cell_info_sheet import add_cell_info_sheet
            page_id = page.section_id if page and page.section_id else "section_0"
            add_cell_info_sheet(wb, hwpx_path, "CellInfo", hide_para_rows, page_id, "")

        # 저장
        wb.save(output_path)

        top_count = len(top_level_tables)
        nested_count = len(nested_tables)
        if inline_nested:
            print(f"  {len(tables)}개 테이블 (최상위: {top_count}, 중첩: {nested_count} -> 인라인)")
        else:
            print(f"  {len(tables)}개 테이블 (최상위: {top_count}, 중첩: {nested_count})")
        return output_path


# ============================================================
# 편의 함수
# ============================================================

def convert_hwpx_to_excel(
    hwpx_path: Union[str, Path],
    output_path: Optional[Union[str, Path]] = None,
    table_index: int = 0,
    include_cell_info: bool = False,
    hide_para_rows: bool = True
) -> Path:
    """
    HWPX 테이블을 Excel로 변환

    Args:
        hwpx_path: HWPX 파일 경로
        output_path: 출력 Excel 경로
        table_index: 테이블 인덱스
        include_cell_info: 셀 상세 정보 시트 포함 여부
        hide_para_rows: para_id 행 숨김 여부

    Returns:
        생성된 Excel 파일 경로
    """
    converter = HwpxToExcel()
    return converter.convert(
        hwpx_path, output_path, table_index,
        include_cell_info, hide_para_rows
    )


# ============================================================
# 테스트
# ============================================================

if __name__ == "__main__":
    import json
    import platform

    # OS에 따라 경로 설정
    if platform.system() == "Windows":
        hwpx_path = r"C:\hwp_xml\table.hwpx"
        output_path = r"C:\hwp_xml\excel\output4.xlsx"
    else:
        hwpx_path = "/mnt/c/hwp_xml/table.hwpx"
        output_path = "/mnt/c/hwp_xml/excel/output.xlsx"

    print(f"입력: {hwpx_path}")
    print(f"출력: {output_path}")
    print("=" * 50)

    converter = HwpxToExcel()

    # 테이블 정보 확인
    tables = converter.table_parser.from_hwpx(hwpx_path)
    if tables:
        table = tables[0]
        print(f"테이블: {table.row_count}행 x {table.col_count}열")
        print(f"크기: {table.width} x {table.height} HWPUNIT")
        print(f"      {Unit.hwpunit_to_mm(table.width):.1f}mm x {Unit.hwpunit_to_mm(table.height):.1f}mm")

    # 페이지 정보 확인
    pages = converter.page_parser.from_hwpx(hwpx_path)
    if pages:
        page = pages[0]
        print(f"\n페이지: {page.page_size.orientation}")
        print(f"크기: {Unit.hwpunit_to_mm(page.page_size.width):.1f}mm x {Unit.hwpunit_to_mm(page.page_size.height):.1f}mm")
        print(f"여백: 좌{Unit.hwpunit_to_mm(page.margin.left):.1f}mm 우{Unit.hwpunit_to_mm(page.margin.right):.1f}mm")

    # 변환 실행
    print("\n변환 중...")
    result = converter.convert(hwpx_path, output_path)
    print(f"완료: {result}")
