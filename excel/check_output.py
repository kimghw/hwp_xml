# -*- coding: utf-8 -*-
"""생성된 Excel 파일 확인"""

from openpyxl import load_workbook

wb = load_workbook('output_with_cellinfo.xlsx')
print(f"시트 목록: {wb.sheetnames}")

# 본문_메타 시트 확인
ws = wb['본문_메타']
print(f"본문_메타 시트: {ws.max_row}행 x {ws.max_column}열")

# 헤더 확인
headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
print(f"헤더: {headers}")

# 첫 데이터 행 확인
print("\n첫 번째 데이터 행:")
for i, h in enumerate(headers[:10], 1):
    val = ws.cell(2, i).value
    print(f"  {h}: {val}")

# 숨김 행 개수
hidden_count = sum(1 for i in range(2, ws.max_row + 1) if ws.row_dimensions[i].hidden)
print(f"\n숨김 행 개수: {hidden_count}")
