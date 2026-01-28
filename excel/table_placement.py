# -*- coding: utf-8 -*-
"""Excel 테이블 배치 모듈"""

import sys
from pathlib import Path
from typing import List, Dict, Tuple

# 프로젝트 루트 경로 설정
_project_root = Path(__file__).parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

from hwpxml.get_table_property import TableProperty
from hwpxml.get_page_property import PageProperty, Unit
from hwpxml.get_cell_detail import CellDetail


class TablePlacer:
    """테이블 배치 클래스"""

    # 엑셀 열 너비 변환 계수
    # HWPUNIT -> 엑셀 열 너비: hwpunit / 632
    HWPUNIT_TO_EXCEL_WIDTH = 632

    # 엑셀 행 높이는 포인트 단위
    # HWPUNIT -> pt: hwpunit / 100
    HWPUNIT_TO_PT = 100

    # 용지 크기 매핑 (mm -> Excel paperSize 코드)
    PAPER_SIZES = {
        (210, 297): 9,    # A4
        (297, 420): 8,    # A3
        (148, 210): 11,   # A5
        (182, 257): 13,   # B5
        (250, 354): 12,   # B4
        (216, 279): 1,    # Letter (8.5 x 11 in)
        (216, 356): 5,    # Legal (8.5 x 14 in)
    }

    # HWP 테두리 타입 -> openpyxl 스타일 매핑
    BORDER_STYLE_MAP = {
        'NONE': None,
        'SOLID': 'thin',
        'DASH': 'dashed',
        'DOT': 'dotted',
        'DASH_DOT': 'dashDot',
        'DASH_DOT_DOT': 'dashDotDot',
        'DOUBLE': 'double',
        'WAVE': 'thin',  # wave는 thin으로 대체
        'THICK': 'medium',
        'THICK_DOUBLE': 'double',
        'THICK_DASH': 'mediumDashed',
        'THICK_DASH_DOT': 'mediumDashDot',
        'THICK_DASH_DOT_DOT': 'mediumDashDotDot',
    }

    def build_unified_column_grid(
        self, tables: List[TableProperty], merge_threshold: int = 100
    ) -> List[int]:
        """
        모든 테이블의 열 경계를 수집하여 통합 열 그리드 생성
        원본 너비 유지, 근접한 경계는 병합

        Args:
            tables: 테이블 목록
            merge_threshold: 이 값 이하로 근접한 경계는 병합 (HWPUNIT, ~0.35mm)

        Returns:
            정렬된 열 경계 리스트 (HWPUNIT 단위)
        """
        if not tables:
            return [0]

        boundaries = set([0])

        for table in tables:
            col_widths = self.get_column_widths(table)
            x = 0
            for width in col_widths:
                x += width
                boundaries.add(x)

        # 정렬 후 근접 경계 병합
        sorted_boundaries = sorted(boundaries)
        if len(sorted_boundaries) <= 1:
            return sorted_boundaries

        merged = [sorted_boundaries[0]]
        for b in sorted_boundaries[1:]:
            if b - merged[-1] <= merge_threshold:
                # 근접하면 큰 값으로 대체 (끝점 우선)
                merged[-1] = b
            else:
                merged.append(b)

        return merged

    def get_table_column_boundaries(self, table: TableProperty) -> List[int]:
        """테이블의 열 경계 계산 (원본 너비)"""
        col_widths = self.get_column_widths(table)

        boundaries = [0]
        x = 0
        for width in col_widths:
            x += width
            boundaries.append(x)
        return boundaries

    def map_columns_to_unified(
        self, table_boundaries: List[int], unified_boundaries: List[int],
        tolerance: int = 200
    ) -> Dict[int, tuple]:
        """
        테이블 원본 열 인덱스 -> 통합 그리드 (시작열, 끝열) 매핑
        근접 매칭 지원

        Args:
            tolerance: 이 값 이하 차이는 같은 경계로 간주 (HWPUNIT, ~0.7mm)

        Returns:
            {원본_col_idx: (unified_start_col, unified_end_col)}
        """
        mapping = {}

        def find_nearest(x: int) -> int:
            """가장 가까운 통합 경계 인덱스 찾기"""
            best_idx = 0
            best_diff = abs(unified_boundaries[0] - x)
            for i, b in enumerate(unified_boundaries):
                diff = abs(b - x)
                if diff < best_diff:
                    best_diff = diff
                    best_idx = i
            return best_idx if best_diff <= tolerance else -1

        for orig_col in range(len(table_boundaries) - 1):
            start_x = table_boundaries[orig_col]
            end_x = table_boundaries[orig_col + 1]

            unified_start = find_nearest(start_x)
            unified_end = find_nearest(end_x)

            if unified_start >= 0 and unified_end >= 0:
                mapping[orig_col] = (unified_start, unified_end)

        return mapping

    def place_table_unified(
        self, ws: Worksheet, table: TableProperty, cell_details: List[CellDetail],
        col_mapping: Dict[int, tuple], unified_col_widths: List[int], start_row: int
    ) -> int:
        """
        테이블을 통합 시트에 배치 (기본 모드)

        Returns:
            사용된 행 수
        """
        # 셀 디테일 맵
        cell_map = {(cd.row, cd.col): cd for cd in cell_details}

        # 행 높이 설정
        row_heights = self.get_row_heights(table)
        for row_idx, height in enumerate(row_heights):
            excel_row = start_row + row_idx
            height_pt = height / self.HWPUNIT_TO_PT
            height_pt = max(height_pt, 10)
            ws.row_dimensions[excel_row].height = height_pt

        # 셀 배치
        for cd in cell_details:
            if cd.col not in col_mapping:
                continue

            unified_start_col, unified_end_col = col_mapping[cd.col]
            excel_row = start_row + cd.row
            excel_col = unified_start_col + 1  # 1-based

            # col_span 고려: 원본 col_span만큼의 통합 열 범위 계산
            orig_end_col = cd.col + cd.col_span - 1
            if orig_end_col in col_mapping:
                _, unified_span_end = col_mapping[orig_end_col]
            else:
                unified_span_end = unified_end_col

            unified_col_span = unified_span_end - unified_start_col

            try:
                excel_cell = ws.cell(row=excel_row, column=excel_col)
                excel_cell.value = cd.text

                # 병합 여부 확인
                is_col_merged = unified_col_span > 1

                # 스타일 적용 - 병합 셀이면 right 테두리 생략
                self._apply_cell_style_single(
                    excel_cell, cd, 0, 1,
                    is_merged=is_col_merged, is_last_col=False
                )

                # 병합 처리
                if cd.row_span > 1 or unified_col_span > 1:
                    try:
                        ws.merge_cells(
                            start_row=excel_row,
                            start_column=excel_col,
                            end_row=excel_row + cd.row_span - 1,
                            end_column=excel_col + unified_col_span - 1
                        )
                    except ValueError:
                        pass

            except Exception:
                pass

        return table.row_count

    def place_table_with_para_split_unified(
        self, ws: Worksheet, table: TableProperty, cell_details: List[CellDetail],
        col_mapping: Dict[int, tuple], unified_col_widths: List[int], start_row: int
    ) -> int:
        """
        테이블을 통합 시트에 배치 (문단별 분할 모드)

        Returns:
            사용된 행 수
        """
        cell_map = {(cd.row, cd.col): cd for cd in cell_details}

        # 각 행별 최대 문단 수 계산
        row_max_paras = {}
        for row_idx in range(table.row_count):
            max_paras = 1
            for col_idx in range(table.col_count):
                if (row_idx, col_idx) in cell_map:
                    cd = cell_map[(row_idx, col_idx)]
                    para_count = len(cd.paragraphs) if cd.paragraphs else 1
                    max_paras = max(max_paras, para_count)
            row_max_paras[row_idx] = max_paras

        # 원본 행 -> 새 행 오프셋
        row_offset_map = {}
        cumulative = 0
        for row_idx in range(table.row_count):
            row_offset_map[row_idx] = cumulative
            cumulative += row_max_paras.get(row_idx, 1)

        total_rows = cumulative

        # 셀 배치
        for cd in cell_details:
            if cd.col not in col_mapping:
                continue

            unified_start_col, unified_end_col = col_mapping[cd.col]
            new_start_row = start_row + row_offset_map[cd.row]
            excel_col = unified_start_col + 1  # 1-based

            para_count = len(cd.paragraphs) if cd.paragraphs else 1

            # col_span 고려
            orig_end_col = cd.col + cd.col_span - 1
            if orig_end_col in col_mapping:
                _, unified_span_end = col_mapping[orig_end_col]
            else:
                unified_span_end = unified_end_col
            unified_col_span = unified_span_end - unified_start_col

            # 새 row_span 계산
            new_row_span = 0
            for r in range(cd.row_span):
                new_row_span += row_max_paras.get(cd.row + r, 1)

            # 병합 여부 미리 계산
            is_col_merged = unified_col_span > 1

            # 문단별 값 설정
            for para_idx in range(para_count):
                new_row = new_start_row + para_idx

                try:
                    excel_cell = ws.cell(row=new_row, column=excel_col)

                    if cd.paragraphs and para_idx < len(cd.paragraphs):
                        excel_cell.value = cd.paragraphs[para_idx].text
                    elif para_idx == 0:
                        excel_cell.value = cd.text

                    self._apply_cell_style_single(
                        excel_cell, cd, para_idx, para_count,
                        is_merged=is_col_merged, is_last_col=False
                    )

                except Exception:
                    pass

            # 병합 처리
            merge_end_row = new_start_row + new_row_span - 1
            merge_end_col = excel_col + unified_col_span - 1

            needs_row_merge = new_row_span > para_count  # 세로 병합 필요
            needs_col_merge = unified_col_span > 1  # 가로 병합 필요

            if needs_row_merge or needs_col_merge:
                if para_count <= 1:
                    # 문단 0-1개: 전체 범위 병합
                    if new_row_span > 1 or unified_col_span > 1:
                        try:
                            ws.merge_cells(
                                start_row=new_start_row,
                                start_column=excel_col,
                                end_row=merge_end_row,
                                end_column=merge_end_col
                            )
                        except ValueError:
                            pass
                else:
                    # 문단 2개 이상
                    if needs_row_merge:
                        # 세로 병합 필요: 각 문단 행은 가로만 병합, 마지막 문단부터 끝까지 세로+가로 병합
                        if needs_col_merge:
                            for p_idx in range(para_count - 1):  # 마지막 문단 제외
                                try:
                                    ws.merge_cells(
                                        start_row=new_start_row + p_idx,
                                        start_column=excel_col,
                                        end_row=new_start_row + p_idx,
                                        end_column=merge_end_col
                                    )
                                except ValueError:
                                    pass
                        # 마지막 문단 행부터 끝까지 세로+가로 병합
                        try:
                            ws.merge_cells(
                                start_row=new_start_row + para_count - 1,
                                start_column=excel_col,
                                end_row=merge_end_row,
                                end_column=merge_end_col
                            )
                        except ValueError:
                            pass
                    else:
                        # 세로 병합 불필요: 각 문단 행 가로만 병합
                        if needs_col_merge:
                            for p_idx in range(para_count):
                                try:
                                    ws.merge_cells(
                                        start_row=new_start_row + p_idx,
                                        start_column=excel_col,
                                        end_row=new_start_row + p_idx,
                                        end_column=merge_end_col
                                    )
                                except ValueError:
                                    pass

                # 병합 영역의 내부 테두리 제거
                self._apply_merged_cell_borders(
                    ws, cd, new_start_row, excel_col, merge_end_row, merge_end_col
                )

        return total_rows

    def apply_page_settings(self, ws: Worksheet, page: PageProperty):
        """페이지 설정 적용 (HWPX 값 동적 매핑)"""
        # 용지 방향
        if page.page_size.orientation == 'landscape':
            ws.page_setup.orientation = 'landscape'
        else:
            ws.page_setup.orientation = 'portrait'

        # 용지 크기 (HWPX에서 가져온 값으로 매핑)
        width_mm = round(Unit.hwpunit_to_mm(page.page_size.width))
        height_mm = round(Unit.hwpunit_to_mm(page.page_size.height))

        # 용지 크기 매핑 (가로/세로 모두 확인)
        paper_size = None
        for (w, h), code in self.PAPER_SIZES.items():
            if (abs(width_mm - w) < 5 and abs(height_mm - h) < 5) or \
               (abs(width_mm - h) < 5 and abs(height_mm - w) < 5):
                paper_size = code
                break

        if paper_size:
            ws.page_setup.paperSize = paper_size
        else:
            # 매핑 안되면 사용자 지정 크기 (인치 단위)
            ws.page_setup.paperWidth = f"{page.page_size.width / Unit.HWPUNIT_PER_INCH:.2f}in"
            ws.page_setup.paperHeight = f"{page.page_size.height / Unit.HWPUNIT_PER_INCH:.2f}in"

        # 여백 (HWPX에서 가져온 값을 인치로 변환)
        ws.page_margins = PageMargins(
            left=page.margin.left / Unit.HWPUNIT_PER_INCH,
            right=page.margin.right / Unit.HWPUNIT_PER_INCH,
            top=page.margin.top / Unit.HWPUNIT_PER_INCH,
            bottom=page.margin.bottom / Unit.HWPUNIT_PER_INCH,
            header=page.margin.header / Unit.HWPUNIT_PER_INCH,
            footer=page.margin.footer / Unit.HWPUNIT_PER_INCH,
        )

    def apply_column_widths(self, ws: Worksheet, table: TableProperty):
        """열 너비 설정"""
        col_widths = self.get_column_widths(table)

        for col_idx, width in enumerate(col_widths, start=1):
            col_letter = get_column_letter(col_idx)
            # HWPUNIT -> 엑셀 열 너비
            excel_width = width / self.HWPUNIT_TO_EXCEL_WIDTH
            # 최소 너비 보장
            excel_width = max(excel_width, 1)
            ws.column_dimensions[col_letter].width = excel_width

    def apply_row_heights(self, ws: Worksheet, table: TableProperty):
        """행 높이 설정"""
        row_heights = self.get_row_heights(table)

        for row_idx, height in enumerate(row_heights, start=1):
            # HWPUNIT -> 포인트
            excel_height = height / self.HWPUNIT_TO_PT
            # 최소 높이 보장
            excel_height = max(excel_height, 10)
            ws.row_dimensions[row_idx].height = excel_height

    def apply_cell_merges(self, ws: Worksheet, table: TableProperty):
        """셀 병합 처리"""
        for row in table.cells:
            for cell in row:
                if cell.col_span > 1 or cell.row_span > 1:
                    start_row = cell.row_index + 1  # 1-based
                    start_col = cell.col_index + 1
                    end_row = start_row + cell.row_span - 1
                    end_col = start_col + cell.col_span - 1

                    start_cell = f"{get_column_letter(start_col)}{start_row}"
                    end_cell = f"{get_column_letter(end_col)}{end_row}"

                    try:
                        ws.merge_cells(f"{start_cell}:{end_cell}")
                    except ValueError:
                        # 이미 병합된 셀이면 무시
                        pass

    def apply_table_with_para_split(
        self, ws: Worksheet, table: TableProperty, cell_details: List[CellDetail]
    ) -> dict:
        """
        문단별 행 분할하여 테이블 생성

        원본 테이블의 각 행을 문단 수에 따라 여러 행으로 분할합니다.
        - 셀에 문단이 3개 있으면 해당 행이 3행으로 분할
        - 같은 행의 다른 셀(문단 1개)은 분할된 행에 맞춰 세로 병합
        - 원본 셀 높이를 분할된 행에 균등 분배 (높이 유지)

        Returns:
            row_offset_map: 원본 행 인덱스 -> 새 행 오프셋 매핑
        """
        # 셀 디테일을 (row, col) -> CellDetail 매핑으로 변환
        cell_map = {}
        for cd in cell_details:
            cell_map[(cd.row, cd.col)] = cd

        # 1. 각 행별 최대 문단 수 및 문단별 높이 계산
        row_max_paras = {}
        row_para_heights = {}  # row_idx -> [높이1, 높이2, ...] (각 문단별 최대 높이)

        for row_idx in range(table.row_count):
            max_paras = 1
            para_heights = []

            for col_idx in range(table.col_count):
                if (row_idx, col_idx) in cell_map:
                    cd = cell_map[(row_idx, col_idx)]
                    para_count = len(cd.paragraphs) if cd.paragraphs else 1
                    max_paras = max(max_paras, para_count)

                    # 각 문단별 높이 수집
                    for p_idx, para in enumerate(cd.paragraphs or []):
                        while len(para_heights) <= p_idx:
                            para_heights.append(0)
                        if para.height > para_heights[p_idx]:
                            para_heights[p_idx] = para.height

            row_max_paras[row_idx] = max_paras
            row_para_heights[row_idx] = para_heights if para_heights else [0]

        # 2. 원본 행 -> 새 행 오프셋 매핑 계산
        row_offset_map = {}
        cumulative = 0
        for row_idx in range(table.row_count):
            row_offset_map[row_idx] = cumulative
            cumulative += row_max_paras.get(row_idx, 1)

        # 3. 분할된 행 높이 설정 (각 문단의 실제 높이 사용)
        for row_idx in range(table.row_count):
            para_count = row_max_paras.get(row_idx, 1)
            para_heights = row_para_heights.get(row_idx, [])
            new_start_row = row_offset_map[row_idx] + 1  # 1-based

            for p_idx in range(para_count):
                # 해당 문단의 높이 사용 (없으면 0)
                height = para_heights[p_idx] if p_idx < len(para_heights) else 0
                if height > 0:
                    height_pt = height / self.HWPUNIT_TO_PT
                    ws.row_dimensions[new_start_row + p_idx].height = height_pt

        # 4. 셀 배치
        for cd in cell_details:
            orig_row = cd.row
            orig_col = cd.col
            new_start_row = row_offset_map[orig_row] + 1  # 1-based
            new_col = orig_col + 1  # 1-based

            para_count = len(cd.paragraphs) if cd.paragraphs else 1

            # 이 셀이 차지하는 새 행 수 계산 (row_span 고려)
            new_row_span = 0
            for r in range(cd.row_span):
                new_row_span += row_max_paras.get(orig_row + r, 1)

            # 병합 여부 미리 계산
            is_col_merged = cd.col_span > 1

            # 문단별로 값 설정
            for para_idx in range(para_count):
                new_row = new_start_row + para_idx

                try:
                    excel_cell = ws.cell(row=new_row, column=new_col)

                    if cd.paragraphs and para_idx < len(cd.paragraphs):
                        para = cd.paragraphs[para_idx]
                        excel_cell.value = para.text
                    elif para_idx == 0:
                        excel_cell.value = cd.text

                    # 스타일 적용 - 병합 셀이면 right 테두리 생략
                    self._apply_cell_style_single(
                        excel_cell, cd, para_idx, para_count,
                        is_merged=is_col_merged, is_last_col=False
                    )

                except Exception:
                    pass

            # 병합 처리
            merge_end_row = new_start_row + new_row_span - 1
            merge_start_col = new_col
            merge_end_col = new_col + cd.col_span - 1

            needs_row_merge = new_row_span > para_count  # 세로 병합 필요
            needs_col_merge = cd.col_span > 1  # 가로 병합 필요

            if needs_row_merge or needs_col_merge:
                if para_count <= 1:
                    # 문단이 0-1개: 전체 범위 병합
                    if new_row_span > 1 or cd.col_span > 1:
                        try:
                            ws.merge_cells(
                                start_row=new_start_row,
                                start_column=merge_start_col,
                                end_row=merge_end_row,
                                end_column=merge_end_col
                            )
                        except ValueError:
                            pass
                else:
                    # 문단이 2개 이상
                    if needs_row_merge:
                        # 세로 병합 필요: 마지막 문단 행부터 끝까지 병합 (가로도 포함)
                        # 먼저 마지막 문단 이전 행들은 가로만 병합
                        if needs_col_merge:
                            for p_idx in range(para_count - 1):  # 마지막 제외
                                try:
                                    ws.merge_cells(
                                        start_row=new_start_row + p_idx,
                                        start_column=merge_start_col,
                                        end_row=new_start_row + p_idx,
                                        end_column=merge_end_col
                                    )
                                except ValueError:
                                    pass
                        # 마지막 문단 행부터 끝까지 세로+가로 병합
                        try:
                            ws.merge_cells(
                                start_row=new_start_row + para_count - 1,
                                start_column=merge_start_col,
                                end_row=merge_end_row,
                                end_column=merge_end_col
                            )
                        except ValueError:
                            pass
                    else:
                        # 세로 병합 불필요: 각 문단 행 가로만 병합
                        if needs_col_merge:
                            for p_idx in range(para_count):
                                try:
                                    ws.merge_cells(
                                        start_row=new_start_row + p_idx,
                                        start_column=merge_start_col,
                                        end_row=new_start_row + p_idx,
                                        end_column=merge_end_col
                                    )
                                except ValueError:
                                    pass

                # 병합 영역의 내부 테두리 제거 (외곽만 유지)
                self._apply_merged_cell_borders(
                    ws, cd, new_start_row, merge_start_col, merge_end_row, merge_end_col
                )

        return row_offset_map

    def get_column_widths(self, table: TableProperty) -> List[int]:
        """각 열의 너비 추출 (셀 너비를 span으로 나눠서 분배)"""
        if table.col_count == 0:
            return []

        col_widths = [0] * table.col_count

        # 모든 셀에서 너비 추출하여 각 열에 분배
        for row in table.cells:
            for cell in row:
                if cell.width and cell.width > 0:
                    # 셀 너비를 span 수로 나눠서 각 열에 분배
                    per_col_width = cell.width // cell.col_span
                    for i in range(cell.col_span):
                        col_idx = cell.col_index + i
                        if col_idx < table.col_count:
                            # 아직 설정 안된 열만 설정
                            if col_widths[col_idx] == 0:
                                col_widths[col_idx] = per_col_width

        # 여전히 0인 열은 테이블 너비 기준 균등 분배
        if table.width:
            default_width = table.width // table.col_count
            for i in range(len(col_widths)):
                if col_widths[i] == 0:
                    col_widths[i] = default_width

        return col_widths

    def get_row_heights(self, table: TableProperty) -> List[int]:
        """각 행의 높이 추출 (셀 높이를 span으로 나눠서 분배)"""
        if table.row_count == 0:
            return []

        row_heights = [0] * table.row_count

        # 모든 셀에서 높이 추출하여 각 행에 분배
        for row in table.cells:
            for cell in row:
                if cell.height and cell.height > 0:
                    # 셀 높이를 span 수로 나눠서 각 행에 분배
                    per_row_height = cell.height // cell.row_span
                    for i in range(cell.row_span):
                        row_idx = cell.row_index + i
                        if row_idx < table.row_count:
                            # 아직 설정 안된 행만 설정
                            if row_heights[row_idx] == 0:
                                row_heights[row_idx] = per_row_height

        # 여전히 0인 행은 테이블 높이 기준 균등 분배
        if table.height:
            default_height = table.height // table.row_count
            for i in range(len(row_heights)):
                if row_heights[i] == 0:
                    row_heights[i] = default_height

        return row_heights

    def _apply_merged_cell_borders(
        self, ws: Worksheet, cd: CellDetail,
        start_row: int, start_col: int, end_row: int, end_col: int
    ):
        """병합된 셀 영역의 테두리 적용 (외곽만 유지, 내부 제거)"""
        # 병합 영역이 1x1이면 적용 불필요
        if start_row == end_row and start_col == end_col:
            return

        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                try:
                    cell = ws.cell(row=r, column=c)
                    # 외곽 테두리만 설정
                    border = Border(
                        left=self._get_border_side(cd.border.left) if c == start_col else Side(),
                        right=self._get_border_side(cd.border.right) if c == end_col else Side(),
                        top=self._get_border_side(cd.border.top) if r == start_row else Side(),
                        bottom=self._get_border_side(cd.border.bottom) if r == end_row else Side(),
                    )
                    cell.border = border
                except:
                    pass

    def _apply_cell_style_single(
        self, excel_cell, cd: CellDetail, para_idx: int, total_paras: int,
        is_merged: bool = False, is_last_col: bool = True
    ):
        """
        단일 셀에 스타일 적용

        Args:
            is_merged: 병합된 셀 여부 (True면 right 테두리 생략)
            is_last_col: 병합 영역의 마지막 열인지 여부
        """
        # 테두리 - 병합 셀이면 right 테두리는 마지막 열에서만 설정
        border = Border(
            left=self._get_border_side(cd.border.left),
            right=self._get_border_side(cd.border.right) if (not is_merged or is_last_col) else Side(),
            top=self._get_border_side(cd.border.top) if para_idx == 0 else Side(),
            bottom=self._get_border_side(cd.border.bottom) if para_idx == total_paras - 1 else Side(),
        )
        excel_cell.border = border

        # 배경색
        bg_color = self._hwp_color_to_rgb(cd.border.bg_color)
        if bg_color and bg_color != 'FFFFFF':
            excel_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        # 폰트 (문단별 폰트 사용, 없으면 셀 기본 폰트)
        para_font = None
        if para_idx < len(cd.paragraphs) and cd.paragraphs[para_idx].font:
            para_font = cd.paragraphs[para_idx].font
        else:
            para_font = cd.font

        font_color = self._hwp_color_to_rgb(para_font.color)
        excel_cell.font = Font(
            name=para_font.name if para_font.name else None,
            size=para_font.size_pt() if para_font.size > 0 else None,
            bold=para_font.bold,
            italic=para_font.italic,
            underline='single' if para_font.underline else None,
            strike=para_font.strikeout,
            color=font_color if font_color else None,
        )

        # 정렬
        h_align_map = {'LEFT': 'left', 'CENTER': 'center', 'RIGHT': 'right', 'JUSTIFY': 'justify'}
        v_align_map = {'TOP': 'top', 'CENTER': 'center', 'BOTTOM': 'bottom', 'BASELINE': 'center'}
        h_align = 'left'
        v_align = 'center'
        if cd.paragraphs:
            h_align = h_align_map.get(cd.paragraphs[0].align_h, 'left')
            v_align = v_align_map.get(cd.paragraphs[0].align_v, 'center')
        excel_cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=True)

    def _get_border_side(self, border_type: str) -> Side:
        """HWP 테두리 타입을 openpyxl Side로 변환"""
        style = self.BORDER_STYLE_MAP.get(border_type, None)
        if style:
            return Side(style=style, color='000000')
        return Side(style=None)

    def _hwp_color_to_rgb(self, color_str: str) -> str:
        """HWP 색상 문자열을 RGB hex로 변환"""
        if not color_str:
            return None
        # #RRGGBB 형식이면 그대로 반환 (# 제거)
        if color_str.startswith('#'):
            return color_str[1:].upper()
        # RGB(r,g,b) 형식 처리
        if color_str.startswith('RGB('):
            try:
                rgb = color_str[4:-1].split(',')
                r, g, b = int(rgb[0]), int(rgb[1]), int(rgb[2])
                return f"{r:02X}{g:02X}{b:02X}"
            except:
                return None
        # 숫자형 색상 (BGR 또는 RGB)
        try:
            val = int(color_str)
            r = val & 0xFF
            g = (val >> 8) & 0xFF
            b = (val >> 16) & 0xFF
            return f"{r:02X}{g:02X}{b:02X}"
        except:
            return None
