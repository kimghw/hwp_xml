# -*- coding: utf-8 -*-
"""Excel 셀 스타일 적용 모듈"""

import sys
from pathlib import Path

# 프로젝트 루트 경로 설정
_project_root = Path(__file__).parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from typing import List

from hwpxml.get_cell_detail import CellDetail


class ExcelStyler:
    """Excel 셀 스타일 적용 클래스"""

    # HWP 테두리 타입 -> openpyxl 스타일 매핑
    BORDER_STYLE_MAP = {
        'NONE': None,
        'SOLID': 'thin',
        'DASH': 'dashed',
        'DOT': 'dotted',
        'DASH_DOT': 'dashDot',
        'DASH_DOT_DOT': 'dashDotDot',
        'DOUBLE': 'double',
        'WAVE': 'thin',  # wave는 thin으로 대체
        'THICK': 'medium',
        'THICK_DOUBLE': 'double',
        'THICK_DASH': 'mediumDashed',
        'THICK_DASH_DOT': 'mediumDashDot',
        'THICK_DASH_DOT_DOT': 'mediumDashDotDot',
    }

    def hwp_color_to_rgb(self, color_str: str) -> str:
        """HWP 색상 문자열을 RGB hex로 변환"""
        if not color_str:
            return None
        # #RRGGBB 형식이면 그대로 반환 (# 제거)
        if color_str.startswith('#'):
            return color_str[1:].upper()
        # RGB(r,g,b) 형식 처리
        if color_str.startswith('RGB('):
            try:
                rgb = color_str[4:-1].split(',')
                r, g, b = int(rgb[0]), int(rgb[1]), int(rgb[2])
                return f"{r:02X}{g:02X}{b:02X}"
            except:
                return None
        # 숫자형 색상 (BGR 또는 RGB)
        try:
            val = int(color_str)
            r = val & 0xFF
            g = (val >> 8) & 0xFF
            b = (val >> 16) & 0xFF
            return f"{r:02X}{g:02X}{b:02X}"
        except:
            return None

    def get_border_side(self, border_type: str) -> Side:
        """HWP 테두리 타입을 openpyxl Side로 변환"""
        style = self.BORDER_STYLE_MAP.get(border_type, None)
        if style:
            return Side(style=style, color='000000')
        return Side(style=None)

    def apply_cell_styles(self, ws: Worksheet, cell_details: List[CellDetail], table_idx: int = 0):
        """셀 스타일 적용 (테두리, 배경색, 텍스트, 폰트)"""
        for cell_detail in cell_details:
            row = cell_detail.row + 1  # 1-based
            col = cell_detail.col + 1

            try:
                excel_cell = ws.cell(row=row, column=col)
            except:
                continue

            # 병합된 셀의 마스터가 아닌 경우 스킵
            # openpyxl에서 MergedCell은 읽기 전용
            if hasattr(excel_cell, 'is_merged') or type(excel_cell).__name__ == 'MergedCell':
                # 병합 영역의 첫 번째 셀인지 확인
                cell_coord = f"{get_column_letter(col)}{row}"
                is_master = True
                for merged_range in ws.merged_cells.ranges:
                    if cell_coord in merged_range and cell_coord != str(merged_range).split(':')[0]:
                        is_master = False
                        break
                if not is_master:
                    continue

            # 1. 텍스트 설정 (하이퍼링크가 아닌 경우에만)
            try:
                if not excel_cell.hyperlink and cell_detail.text:
                    excel_cell.value = cell_detail.text
            except AttributeError:
                # MergedCell인 경우 무시
                pass

            # 2. 테두리 설정
            border = Border(
                left=self.get_border_side(cell_detail.border.left),
                right=self.get_border_side(cell_detail.border.right),
                top=self.get_border_side(cell_detail.border.top),
                bottom=self.get_border_side(cell_detail.border.bottom),
            )
            excel_cell.border = border

            # 3. 배경색 설정
            bg_color = self.hwp_color_to_rgb(cell_detail.border.bg_color)
            if bg_color and bg_color != 'FFFFFF':
                excel_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

            # 4. 폰트 설정
            font_color = self.hwp_color_to_rgb(cell_detail.font.color)
            excel_cell.font = Font(
                name=cell_detail.font.name if cell_detail.font.name else None,
                size=cell_detail.font.size_pt() if cell_detail.font.size > 0 else None,
                bold=cell_detail.font.bold,
                italic=cell_detail.font.italic,
                underline='single' if cell_detail.font.underline else None,
                strike=cell_detail.font.strikeout,
                color=font_color if font_color else None,
            )

            # 5. 정렬 설정
            h_align_map = {'LEFT': 'left', 'CENTER': 'center', 'RIGHT': 'right', 'JUSTIFY': 'justify'}
            v_align_map = {'TOP': 'top', 'CENTER': 'center', 'BOTTOM': 'bottom', 'BASELINE': 'center'}

            h_align = 'left'
            v_align = 'center'
            if cell_detail.paragraphs:
                h_align = h_align_map.get(cell_detail.paragraphs[0].align_h, 'left')
                v_align = v_align_map.get(cell_detail.paragraphs[0].align_v, 'center')

            excel_cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=True)

            # 6. 병합된 셀의 나머지 영역에도 테두리 적용
            if cell_detail.row_span > 1 or cell_detail.col_span > 1:
                for r in range(cell_detail.row_span):
                    for c in range(cell_detail.col_span):
                        if r == 0 and c == 0:
                            continue  # 첫 셀은 이미 처리됨
                        try:
                            merged_cell = ws.cell(row=row + r, column=col + c)
                            # 병합된 영역의 테두리만 설정
                            merged_border = Border(
                                left=self.get_border_side(cell_detail.border.left) if c == 0 else Side(),
                                right=self.get_border_side(cell_detail.border.right) if c == cell_detail.col_span - 1 else Side(),
                                top=self.get_border_side(cell_detail.border.top) if r == 0 else Side(),
                                bottom=self.get_border_side(cell_detail.border.bottom) if r == cell_detail.row_span - 1 else Side(),
                            )
                            merged_cell.border = merged_border
                        except:
                            pass

    def apply_cell_style_single(
        self, excel_cell, cd: CellDetail, para_idx: int, total_paras: int,
        is_merged: bool = False, is_last_col: bool = True
    ):
        """
        단일 셀에 스타일 적용

        Args:
            is_merged: 병합된 셀 여부 (True면 right 테두리 생략)
            is_last_col: 병합 영역의 마지막 열인지 여부
        """
        # 테두리 - 병합 셀이면 right 테두리는 마지막 열에서만 설정
        border = Border(
            left=self.get_border_side(cd.border.left),
            right=self.get_border_side(cd.border.right) if (not is_merged or is_last_col) else Side(),
            top=self.get_border_side(cd.border.top) if para_idx == 0 else Side(),
            bottom=self.get_border_side(cd.border.bottom) if para_idx == total_paras - 1 else Side(),
        )
        excel_cell.border = border

        # 배경색
        bg_color = self.hwp_color_to_rgb(cd.border.bg_color)
        if bg_color and bg_color != 'FFFFFF':
            excel_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        # 폰트 (문단별 폰트 사용, 없으면 셀 기본 폰트)
        para_font = None
        if para_idx < len(cd.paragraphs) and cd.paragraphs[para_idx].font:
            para_font = cd.paragraphs[para_idx].font
        else:
            para_font = cd.font

        font_color = self.hwp_color_to_rgb(para_font.color)
        excel_cell.font = Font(
            name=para_font.name if para_font.name else None,
            size=para_font.size_pt() if para_font.size > 0 else None,
            bold=para_font.bold,
            italic=para_font.italic,
            underline='single' if para_font.underline else None,
            strike=para_font.strikeout,
            color=font_color if font_color else None,
        )

        # 정렬
        h_align_map = {'LEFT': 'left', 'CENTER': 'center', 'RIGHT': 'right', 'JUSTIFY': 'justify'}
        v_align_map = {'TOP': 'top', 'CENTER': 'center', 'BOTTOM': 'bottom', 'BASELINE': 'center'}
        h_align = 'left'
        v_align = 'center'
        if cd.paragraphs:
            h_align = h_align_map.get(cd.paragraphs[0].align_h, 'left')
            v_align = v_align_map.get(cd.paragraphs[0].align_v, 'center')
        excel_cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=True)

    def apply_merged_cell_borders(
        self, ws, cd: CellDetail,
        start_row: int, start_col: int, end_row: int, end_col: int
    ):
        """병합된 셀 영역의 테두리 적용 (외곽만 유지, 내부 제거)"""
        # 병합 영역이 1x1이면 적용 불필요
        if start_row == end_row and start_col == end_col:
            return

        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                try:
                    cell = ws.cell(row=r, column=c)
                    # 외곽 테두리만 설정
                    border = Border(
                        left=self.get_border_side(cd.border.left) if c == start_col else Side(),
                        right=self.get_border_side(cd.border.right) if c == end_col else Side(),
                        top=self.get_border_side(cd.border.top) if r == start_row else Side(),
                        bottom=self.get_border_side(cd.border.bottom) if r == end_row else Side(),
                    )
                    cell.border = border
                except:
                    pass

    def apply_nested_cell_style(
        self, excel_cell, cd: CellDetail, para_idx: int, total_paras: int
    ):
        """
        nested 테이블 셀에 스타일 적용 (점선 테두리)
        """
        # 점선 테두리 스타일
        dashed_side = Side(style='dashed', color='808080')

        border = Border(
            left=dashed_side,
            right=dashed_side,
            top=dashed_side if para_idx == 0 else Side(),
            bottom=dashed_side if para_idx == total_paras - 1 else Side(),
        )
        excel_cell.border = border

        # 배경색
        bg_color = self.hwp_color_to_rgb(cd.border.bg_color)
        if bg_color and bg_color != 'FFFFFF':
            excel_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        # 폰트
        para_font = cd.font
        if para_idx < len(cd.paragraphs) and cd.paragraphs[para_idx].font:
            para_font = cd.paragraphs[para_idx].font

        font_color = self.hwp_color_to_rgb(para_font.color)
        excel_cell.font = Font(
            name=para_font.name if para_font.name else None,
            size=para_font.size_pt() if para_font.size > 0 else None,
            bold=para_font.bold,
            italic=para_font.italic,
            underline='single' if para_font.underline else None,
            strike=para_font.strikeout,
            color=font_color if font_color else None,
        )

        # 정렬
        h_align_map = {'LEFT': 'left', 'CENTER': 'center', 'RIGHT': 'right', 'JUSTIFY': 'justify'}
        v_align_map = {'TOP': 'top', 'CENTER': 'center', 'BOTTOM': 'bottom', 'BASELINE': 'center'}
        h_align = 'left'
        v_align = 'center'
        if cd.paragraphs:
            h_align = h_align_map.get(cd.paragraphs[0].align_h, 'left')
            v_align = v_align_map.get(cd.paragraphs[0].align_v, 'center')
        excel_cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=True)

    def apply_nested_merged_borders(
        self, ws, cd: CellDetail,
        start_row: int, start_col: int, end_row: int, end_col: int
    ):
        """nested 테이블의 병합 셀 영역에 점선 테두리 적용"""
        if start_row == end_row and start_col == end_col:
            return

        dashed_side = Side(style='dashed', color='808080')

        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                try:
                    cell = ws.cell(row=r, column=c)
                    border = Border(
                        left=dashed_side if c == start_col else Side(),
                        right=dashed_side if c == end_col else Side(),
                        top=dashed_side if r == start_row else Side(),
                        bottom=dashed_side if r == end_row else Side(),
                    )
                    cell.border = border
                except:
                    pass
