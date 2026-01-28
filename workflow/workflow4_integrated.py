# -*- coding: utf-8 -*-
"""
Workflow 4: 통합 Excel 생성

workflow1 + workflow2 + workflow3 조합하여 완전한 메타데이터가 포함된 Excel 생성

1. HWP 열기
2. workflow1 실행 → `파일_meta.yaml`
3. workflow2 실행 → `파일_para.yaml`
4. workflow3 실행 → `파일.xlsx`
5. 정리 (임시 파일 삭제)

출력 파일:
- {base}_meta.yaml  : 테이블/셀 메타데이터
- {base}_para.yaml  : 문단 스타일 정보
- {base}.xlsx       : Excel 변환 결과
"""

# ============================================================
# 설정
# ============================================================
# 기본 HWP 파일 경로 (명령줄 인자 없을 때 사용)
DEFAULT_HWP_PATH = None  # 스크립트와 동일한 이름의 .hwp 파일 (main에서 설정)
# ============================================================

import sys
import os
import tempfile
from pathlib import Path

# 프로젝트 루트 경로 설정
_project_root = Path(__file__).parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

try:
    from config import WIN32HWP_DIR
    if str(WIN32HWP_DIR) not in sys.path:
        sys.path.insert(0, str(WIN32HWP_DIR))
except ImportError:
    win32hwp_dir = os.environ.get('WIN32HWP_DIR', r'C:\win32hwp')
    if win32hwp_dir not in sys.path:
        sys.path.insert(0, win32hwp_dir)

from win32.hwp_file_manager import get_hwp_instance, create_hwp_instance, get_active_filepath, open_file_dialog, save_hwp, open_hwp


class Workflow4:
    """통합 워크플로우 실행기"""

    def __init__(self):
        self.hwp = None
        self.hwp_created = False  # 새로 생성했는지 여부
        self.filepath = None
        self.temp_hwpx = None
        # 메타데이터 저장용
        self.cell_positions = None
        self.field_names = None
        self.para_styles = None

    def _get_hwp(self):
        """한글 인스턴스 가져오기 (기존 연결 또는 새로 생성)"""
        self.hwp = get_hwp_instance()
        if self.hwp:
            print("기존 한글 인스턴스 연결")
            self.hwp_created = False
            try:
                self.hwp.RegisterModule("FilePathCheckDLL", "SecurityModule")
            except:
                pass
        else:
            print("새 한글 인스턴스 생성")
            self.hwp = create_hwp_instance(visible=True)
            self.hwp_created = True
        return self.hwp

    def _open_file(self, filepath: str = None) -> str:
        """HWP 파일 열기"""
        if filepath:
            self.filepath = filepath
        else:
            # 1. 열려있는 문서 확인
            self.filepath = get_active_filepath(self.hwp)
            if self.filepath:
                print(f"열린 문서 사용: {self.filepath}")
                return self.filepath

            # 2. 파일 선택 대화상자
            print("파일을 선택하세요...")
            self.filepath = open_file_dialog()
            if not self.filepath:
                raise ValueError("파일이 선택되지 않았습니다.")

        # 파일 열기
        if not get_active_filepath(self.hwp) == self.filepath:
            open_hwp(self.hwp,self.filepath)
        print(f"파일 열림: {self.filepath}")
        return self.filepath

    def _save_as_hwpx(self) -> str:
        """임시 HWPX로 저장"""
        temp_dir = tempfile.gettempdir()
        self.temp_hwpx = os.path.join(temp_dir, "workflow4_temp.hwpx")

        # 기존 임시 파일 삭제
        if os.path.exists(self.temp_hwpx):
            try:
                os.remove(self.temp_hwpx)
            except:
                pass

        self.hwp.HAction.GetDefault("FileSaveAs_S", self.hwp.HParameterSet.HFileOpenSave.HSet)
        self.hwp.HParameterSet.HFileOpenSave.filename = self.temp_hwpx
        self.hwp.HParameterSet.HFileOpenSave.Format = "HWPX"
        self.hwp.HAction.Execute("FileSaveAs_S", self.hwp.HParameterSet.HFileOpenSave.HSet)

        # 문서 닫고 파일 잠금 해제
        self.hwp.Clear(1)

        print(f"임시 HWPX 저장: {self.temp_hwpx}")
        return self.temp_hwpx

    def _run_workflow1(self, base_path: str) -> str:
        """
        Workflow 1: 테이블 메타데이터 추출
        → {base}_meta.yaml
        """
        print("\n" + "=" * 60)
        print("Workflow 1: 테이블 메타데이터 추출")
        print("=" * 60)

        from win32.insert_table_field import InsertTableField

        inserter = InsertTableField(self.hwp)

        # 필드명 삽입
        print("필드명 삽입 중...")
        count = inserter.insert_field_to_xml(self.temp_hwpx)
        print(f"  {count}개 테이블 처리")

        if count == 0:
            print("  테이블 없음, 건너뜀")
            return None

        # 수정된 HWPX 열기
        open_hwp(self.hwp,self.temp_hwpx)

        # 캡션 삽입
        print("캡션 삽입 중...")
        table_infos = inserter.collect_table_list_ids()
        caption_count = inserter.insert_caption_text(table_infos)
        print(f"  {caption_count}개 캡션 삽입")

        # 메타데이터 추출
        from win32.extract_cell_meta import ExtractCellMeta

        meta_yaml = base_path + "_meta.yaml"
        extractor = ExtractCellMeta(self.hwp)

        # HWPX 다시 저장 (캡션 포함)
        self.hwp.HAction.GetDefault("FileSaveAs_S", self.hwp.HParameterSet.HFileOpenSave.HSet)
        self.hwp.HParameterSet.HFileOpenSave.filename = self.temp_hwpx
        self.hwp.HParameterSet.HFileOpenSave.Format = "HWPX"
        self.hwp.HAction.Execute("FileSaveAs_S", self.hwp.HParameterSet.HFileOpenSave.HSet)

        # COM API로 셀 위치 추출 + HWPX에서 필드명 추출 → YAML 생성
        self.cell_positions = extractor._extract_cell_positions()
        self.field_names = extractor._extract_field_names_from_hwpx(self.temp_hwpx)
        yaml_content = extractor._merge_to_yaml(self.cell_positions, self.field_names)

        with open(meta_yaml, 'w', encoding='utf-8') as f:
            f.write(yaml_content)

        print(f"메타데이터 저장: {meta_yaml}")

        # 필드명 삭제 (HWPX XML 직접 수정)
        self._clear_field_names_in_hwpx(self.temp_hwpx)

        return meta_yaml

    def _clear_field_names_in_hwpx(self, hwpx_path: str):
        """HWPX에서 tc.name 속성만 삭제 (ZIP 직접 수정)"""
        import zipfile
        import shutil
        import xml.etree.ElementTree as ET

        extract_dir = tempfile.mkdtemp()
        total_cleared = 0

        try:
            with zipfile.ZipFile(hwpx_path, 'r') as zf:
                zf.extractall(extract_dir)

            contents_dir = os.path.join(extract_dir, 'Contents')
            section_files = sorted([
                f for f in os.listdir(contents_dir)
                if f.startswith('section') and f.endswith('.xml')
            ])

            for section_file in section_files:
                section_path = os.path.join(contents_dir, section_file)
                tree = ET.parse(section_path)
                root = tree.getroot()

                # 모든 tc 태그에서 name 속성 제거
                for tc in root.iter():
                    if tc.tag.endswith('}tc'):
                        if 'name' in tc.attrib:
                            del tc.attrib['name']
                            total_cleared += 1

                tree.write(section_path, encoding='utf-8', xml_declaration=True)

            # 수정된 HWPX 다시 압축
            with zipfile.ZipFile(hwpx_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for root_dir, dirs, files in os.walk(extract_dir):
                    for file in files:
                        file_path = os.path.join(root_dir, file)
                        arcname = os.path.relpath(file_path, extract_dir)
                        zf.write(file_path, arcname)

            print(f"필드명 삭제: {total_cleared}개 셀")

        finally:
            shutil.rmtree(extract_dir, ignore_errors=True)

    def _run_workflow2(self, base_path: str) -> str:
        """
        Workflow 2: 문단 스타일 추출
        → {base}_para.yaml
        """
        print("\n" + "=" * 60)
        print("Workflow 2: 문단 스타일 추출")
        print("=" * 60)

        from win32.get_para_style import GetParaStyle

        getter = GetParaStyle(self.hwp)
        self.para_styles = getter.get_all_para_styles()

        para_yaml = base_path + "_para.yaml"
        with open(para_yaml, 'w', encoding='utf-8') as f:
            f.write(getter.to_yaml(self.para_styles))

        print(f"문단 스타일 저장: {para_yaml}")
        return para_yaml

    def _run_workflow3(self, base_path: str, split_by_para: bool = True) -> str:
        """
        Workflow 3: HWPX → Excel 변환 + 메타데이터 시트 추가
        → {base}.xlsx
        """
        print("\n" + "=" * 60)
        print("Workflow 3: Excel 변환")
        print("=" * 60)

        from excel.hwpx_to_excel import HwpxToExcel
        from openpyxl import load_workbook

        excel_path = base_path + ".xlsx"
        converter = HwpxToExcel()
        converter.convert_all(
            self.temp_hwpx,
            excel_path,
            include_cell_info=False,
            split_by_para=split_by_para
        )

        # 메타데이터 시트 추가
        print("메타데이터 시트 추가 중...")
        wb = load_workbook(excel_path)
        self._add_meta_sheet(wb)
        self._add_para_sheet(wb)
        wb.save(excel_path)

        print(f"Excel 저장: {excel_path}")
        return excel_path

    def _add_meta_sheet(self, wb):
        """테이블/셀 메타데이터 시트 추가"""
        if not self.field_names:
            return

        ws = wb.create_sheet(title="meta")

        # 헤더
        headers = ["tbl_idx", "table_id", "type", "size", "list_range", "caption_list_id", "caption",
                   "row", "col", "row_span", "col_span", "list_id"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        row_num = 2
        for tbl_idx, tbl_data in enumerate(self.field_names):
            # 테이블 정보
            table_id = tbl_data.get('table_id', '')
            tbl_type = 'parent'
            parent_tbl = None

            # 첫 셀에서 타입 파싱
            if tbl_data.get('cells'):
                first_cell = tbl_data['cells'][0]
                if first_cell.get('field_name'):
                    try:
                        import json
                        fd = json.loads(first_cell['field_name'])
                        tbl_type = fd.get('type', 'parent')
                        parent_tbl = fd.get('parentTbl')
                    except:
                        pass

            size = f"{tbl_data.get('row_count', 0)}x{tbl_data.get('col_count', 0)}"

            # list_range 계산
            com_cells = {}
            if tbl_idx < len(self.cell_positions):
                com_cells = self.cell_positions[tbl_idx].get('cells', {})

            list_range = ""
            caption_list_id = ""
            if com_cells:
                list_ids = [v[0] for v in com_cells.values()]
                min_id, max_id = min(list_ids), max(list_ids)
                list_range = f"{min_id}-{max_id}"
                if tbl_type == 'parent':
                    caption_list_id = min_id - 1
                elif tbl_data.get('caption'):
                    caption_list_id = min_id - 1

            caption = tbl_data.get('caption', '')

            # 각 셀 정보
            for cell in tbl_data.get('cells', []):
                r, c = cell['row'], cell['col']
                cell_list_id = ""
                if (r, c) in com_cells:
                    cell_list_id = com_cells[(r, c)][0]

                ws.cell(row=row_num, column=1, value=tbl_idx)
                ws.cell(row=row_num, column=2, value=table_id)
                ws.cell(row=row_num, column=3, value=tbl_type)
                ws.cell(row=row_num, column=4, value=size)
                ws.cell(row=row_num, column=5, value=list_range)
                ws.cell(row=row_num, column=6, value=caption_list_id)
                ws.cell(row=row_num, column=7, value=caption if r == 0 and c == 0 else "")
                ws.cell(row=row_num, column=8, value=r)
                ws.cell(row=row_num, column=9, value=c)
                ws.cell(row=row_num, column=10, value=cell['row_span'])
                ws.cell(row=row_num, column=11, value=cell['col_span'])
                ws.cell(row=row_num, column=12, value=cell_list_id)
                row_num += 1

        # 열 너비
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['G'].width = 30

        print(f"  meta 시트: {row_num - 2}개 셀")

    def _add_para_sheet(self, wb):
        """문단 스타일 시트 추가"""
        if not self.para_styles:
            return

        ws = wb.create_sheet(title="para")

        # 헤더
        headers = ["list_id", "para_id", "text", "line_count",
                   "start_char", "next_line_char", "end_char",
                   "style_name", "align", "indent", "margin_left", "margin_right",
                   "line_spacing", "line_spacing_type", "space_before", "space_after",
                   "font_name", "font_size", "bold", "italic", "underline", "strikeout",
                   "text_color", "highlight_color"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_num, ps in enumerate(self.para_styles, 2):
            ws.cell(row=row_num, column=1, value=ps.list_id)
            ws.cell(row=row_num, column=2, value=ps.para_id)
            # 텍스트 길이 제한 (Excel 셀 최대 32767자)
            text = ps.text[:1000] if ps.text else ""
            ws.cell(row=row_num, column=3, value=text)
            ws.cell(row=row_num, column=4, value=ps.line_count)
            ws.cell(row=row_num, column=5, value=ps.start_char_id)
            ws.cell(row=row_num, column=6, value=ps.next_line_char_id)
            ws.cell(row=row_num, column=7, value=ps.end_char_id)
            ws.cell(row=row_num, column=8, value=ps.style_name)
            ws.cell(row=row_num, column=9, value=ps.align)
            ws.cell(row=row_num, column=10, value=ps.indent)
            ws.cell(row=row_num, column=11, value=ps.margin_left)
            ws.cell(row=row_num, column=12, value=ps.margin_right)
            ws.cell(row=row_num, column=13, value=ps.line_spacing)
            ws.cell(row=row_num, column=14, value=ps.line_spacing_type)
            ws.cell(row=row_num, column=15, value=ps.space_before)
            ws.cell(row=row_num, column=16, value=ps.space_after)

            # 글자 스타일
            if ps.char_style:
                cs = ps.char_style
                ws.cell(row=row_num, column=17, value=cs.font_name)
                ws.cell(row=row_num, column=18, value=cs.font_size)
                ws.cell(row=row_num, column=19, value=cs.bold)
                ws.cell(row=row_num, column=20, value=cs.italic)
                ws.cell(row=row_num, column=21, value=cs.underline)
                ws.cell(row=row_num, column=22, value=cs.strikeout)
                ws.cell(row=row_num, column=23, value=cs.text_color)
                ws.cell(row=row_num, column=24, value=cs.highlight_color)

        # 열 너비
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['Q'].width = 15

        print(f"  para 시트: {len(self.para_styles)}개 문단")

    def _cleanup(self):
        """임시 파일 삭제"""
        print("\n" + "-" * 60)
        print("정리 중...")

        if self.temp_hwpx and os.path.exists(self.temp_hwpx):
            try:
                os.remove(self.temp_hwpx)
                print(f"  삭제: {self.temp_hwpx}")
            except:
                pass

        # clean 파일도 삭제
        clean_hwpx = self.temp_hwpx.replace('.hwpx', '_clean.hwpx') if self.temp_hwpx else None
        if clean_hwpx and os.path.exists(clean_hwpx):
            try:
                os.remove(clean_hwpx)
                print(f"  삭제: {clean_hwpx}")
            except:
                pass

    def _should_close(self) -> bool:
        """종료 여부 판단 (새로 생성한 경우만 닫음)"""
        return self.hwp_created

    def _create_data_dir(self) -> str:
        """data/파일명/ 폴더 생성"""
        if not self.filepath:
            raise ValueError("파일 경로가 설정되지 않았습니다.")

        file_dir = os.path.dirname(self.filepath)
        file_name = os.path.splitext(os.path.basename(self.filepath))[0]

        data_dir = os.path.join(file_dir, 'data', file_name)
        os.makedirs(data_dir, exist_ok=True)

        print(f"결과 폴더: {data_dir}")
        return data_dir

    def run(self, filepath: str = None, split_by_para: bool = True) -> dict:
        """
        통합 워크플로우 실행

        Args:
            filepath: HWP 파일 경로 (없으면 열린 문서 또는 대화상자)
            split_by_para: 문단별 행 분할 여부

        Returns:
            생성된 파일 경로 dict
        """
        print("=" * 60)
        print("Workflow 4: 통합 Excel 생성")
        print("=" * 60)

        results = {
            'meta_yaml': None,
            'para_yaml': None,
            'excel': None,
        }

        try:
            # 1. 한글 인스턴스 가져오기
            self._get_hwp()

            # 2. 파일 열기
            self._open_file(filepath)

            # 기본 경로: data/파일명/파일명
            data_dir = self._create_data_dir()
            file_name = os.path.splitext(os.path.basename(self.filepath))[0]
            base_path = os.path.join(data_dir, file_name)

            # 3. HWPX로 변환
            self._save_as_hwpx()

            # 4. Workflow 1: 메타데이터 추출
            results['meta_yaml'] = self._run_workflow1(base_path)

            # 5. Workflow 2: 문단 스타일 추출
            # HWPX 다시 열기 (workflow1에서 수정되었을 수 있음)
            open_hwp(self.hwp,self.temp_hwpx)
            results['para_yaml'] = self._run_workflow2(base_path)

            # 6. Workflow 3: Excel 변환
            results['excel'] = self._run_workflow3(base_path, split_by_para)

            # 7. 정리
            self._cleanup()

            # 8. 완료
            print("\n" + "=" * 60)
            print("완료!")
            print("=" * 60)
            print(f"  메타데이터: {results['meta_yaml']}")
            print(f"  문단스타일: {results['para_yaml']}")
            print(f"  Excel:     {results['excel']}")

            # 9. 종료 (새로 생성한 경우만)
            if self._should_close():
                try:
                    self.hwp.Quit()
                    print("한글 종료됨")
                except:
                    pass

        except Exception as e:
            print(f"\n오류: {e}")
            import traceback
            traceback.print_exc()

        return results


def main():
    """메인 함수"""
    filepath = None

    # 명령줄 인자 확인
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
        # 상대 경로면 절대 경로로 변환
        if not os.path.isabs(filepath):
            filepath = os.path.abspath(filepath)
    else:
        # 스크립트와 동일한 이름의 .hwp 파일을 기본 경로로
        script_hwp = str(Path(__file__).with_suffix('.hwp'))
        if os.path.exists(script_hwp):
            filepath = script_hwp

    if filepath and not os.path.exists(filepath):
        print(f"파일을 찾을 수 없습니다: {filepath}")
        filepath = None

    workflow = Workflow4()
    workflow.run(filepath)


if __name__ == "__main__":
    main()
