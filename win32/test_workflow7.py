"""
Workflow 7 테스트: Excel meta 시트의 para_text를 HWP 필드에 반영
tc.name 방식 사용 (PutFieldText)
"""
import openpyxl
from hwp_file_manager import get_hwp_instance, create_hwp_instance

def test_workflow7_by_field_name():
    """field_name 방식으로 HWP 필드값 업데이트"""

    # Excel 파일 열기
    excel_path = r'C:\hwp_xml\data\test\test_by_bookmark.xlsx'
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['meta']

    print(f"Excel: {excel_path}")
    print(f"Meta rows: {ws.max_row}")

    # 컬럼 인덱스 (1-based)
    COL_TBL_IDX = 1
    COL_ROW = 3
    COL_COL = 4
    COL_FIELD_NAME = 9
    COL_PARA_TEXT = 11

    # 셀별 문단 합치기
    cell_dict = {}   # cellKey -> 합쳐진 텍스트
    field_dict = {}  # cellKey -> field_name

    for row in range(2, ws.max_row + 1):
        tbl_idx = ws.cell(row, COL_TBL_IDX).value
        cell_row = ws.cell(row, COL_ROW).value
        cell_col = ws.cell(row, COL_COL).value
        para_text = ws.cell(row, COL_PARA_TEXT).value or ""
        field_name = (ws.cell(row, COL_FIELD_NAME).value or "").strip()

        cell_key = f"{tbl_idx}_{cell_row}_{cell_col}"

        # 문단 합치기 (빈 문단도 빈 줄로 유지)
        if cell_key in cell_dict:
            cell_dict[cell_key] = cell_dict[cell_key] + "\n" + str(para_text)
        else:
            cell_dict[cell_key] = str(para_text)

        # field_name 저장 (첫 번째만)
        if field_name and cell_key not in field_dict:
            field_dict[cell_key] = field_name

    print(f"\nfield_name cells: {len(field_dict)}")

    # HWP 인스턴스 가져오기 (이미 열린 문서 사용)
    hwp = get_hwp_instance()
    if hwp is None:
        print("HWP not open. Please open HWP file first.")
        return

    # field_name으로 HWP에 값 입력
    print("\nUpdating fields...")
    update_count = 0
    skip_count = 0

    for cell_key, field_name in field_dict.items():
        combined_text = cell_dict[cell_key]

        if hwp.FieldExist(field_name):
            hwp.PutFieldText(field_name, combined_text)
            update_count += 1
            print(f"  [OK] {field_name[:50]}... -> {combined_text[:20]}...")
        else:
            skip_count += 1
            if skip_count <= 3:
                print(f"  [SKIP] Field not found: {field_name[:50]}...")

    print(f"\nDone: {update_count} updated, {skip_count} skipped")

    wb.close()

if __name__ == "__main__":
    test_workflow7_by_field_name()
